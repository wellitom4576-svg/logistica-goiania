"""
SKW COLETAS - Sistema de Roteirizacao e Coletas
Goiania/GO - Roteirizacao profissional com OR-Tools, motoristas,
bairros coloridos, e otimizacao avancada.
"""

import streamlit as st
import folium
import json
import pandas as pd
import numpy as np
from streamlit_folium import st_folium
from folium.plugins import Draw, MarkerCluster
from shapely.geometry import shape, Point, Polygon, MultiPolygon, MultiPoint
from shapely.ops import unary_union
from geopy.geocoders import Nominatim
from geopy.distance import geodesic
from ortools.constraint_solver import routing_enums_pb2, pywrapcp
import branca.colormap as cm
import io
import zipfile
import colorsys
import os
import math
from datetime import datetime, timedelta
import urllib.parse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Config ──────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="SKW COLETAS",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="expanded",
)

GOIANIA_CENTER = [-16.6869, -49.2648]
DATA_DIR = os.path.join(os.path.dirname(__file__), "data")

CORES_MOTORISTAS = [
    "#e74c3c", "#2ecc71", "#3498db", "#f39c12", "#9b59b6",
    "#1abc9c", "#e67e22", "#e84393", "#00b894", "#6c5ce7",
]
NOMES_CORES = [
    "Vermelho", "Verde", "Azul", "Laranja", "Roxo",
    "Turquesa", "Laranja Esc.", "Rosa", "Menta", "Violeta",
]
ICONES_MOTORISTAS = [
    "🔴", "🟢", "🔵", "🟠", "🟣", "🩵", "🟧", "🩷", "🟩", "🟪",
]
FOLIUM_COLORS = [
    "red", "green", "blue", "orange", "purple",
    "cadetblue", "darkred", "pink", "darkgreen", "darkpurple",
]

# 20 cores para bairros (bem distintas)
CORES_BAIRROS = [
    "#e74c3c", "#2ecc71", "#3498db", "#f39c12", "#9b59b6",
    "#1abc9c", "#e67e22", "#e84393", "#00b894", "#6c5ce7",
    "#fd79a8", "#00cec9", "#fdcb6e", "#6ab04c", "#eb4d4b",
    "#22a6b3", "#f0932b", "#7ed6df", "#c44569", "#574b90",
]

TEMPO_MEDIO_COLETA_MIN = 8  # minutos por parada
VELOCIDADE_MEDIA_KMH = 30   # km/h em cidade
STATUS_PONTO = ["pendente", "coletado", "falhou", "reagendado"]
STATUS_ICONE = {"pendente": "⏳", "coletado": "✅", "falhou": "❌", "reagendado": "🔄"}


# ── Session State ───────────────────────────────────────────────────────────
def init_state():
    defaults = {
        "pontos_coleta": [],
        "rota_otimizada": None,
        "poligonos": [],
        "base_location": GOIANIA_CENTER,
        "motoristas": [],
        "atribuicao_motorista": {},
        "regioes_motoristas": {},
        "bairros_selecionados": {},  # {nome_bairro: cor_idx}
        "geojson_bairros_data": None,
        "tempo_coleta_min": TEMPO_MEDIO_COLETA_MIN,
        "velocidade_media": VELOCIDADE_MEDIA_KMH,
        "hora_inicio": "07:30",
        "status_pontos": {},          # {p_idx: "pendente"|"coletado"|"falhou"|"reagendado"}
        "num_pacotes": {},            # {p_idx: int}
        "capacidade_veiculos": {},    # {m_idx: int}  (0 = sem limite)
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

init_state()


# ── Funcoes Utilitarias ────────────────────────────────────────────────────
def gerar_cores(n):
    cores = []
    for i in range(n):
        hue = i / max(n, 1)
        r, g, b = colorsys.hsv_to_rgb(hue, 0.7, 0.9)
        cores.append(f"#{int(r*255):02x}{int(g*255):02x}{int(b*255):02x}")
    return cores


def calcular_matriz_distancias(pontos):
    n = len(pontos)
    matriz = np.zeros((n, n))
    for i in range(n):
        for j in range(i + 1, n):
            dist = geodesic(
                (pontos[i]["lat"], pontos[i]["lon"]),
                (pontos[j]["lat"], pontos[j]["lon"]),
            ).meters
            matriz[i][j] = int(dist)
            matriz[j][i] = int(dist)
    return matriz


def gerar_link_google_maps(paradas):
    """Gera link do Google Maps com waypoints para navegacao."""
    if len(paradas) < 2:
        return ""
    origin = f"{paradas[0]['lat']},{paradas[0]['lon']}"
    dest = f"{paradas[-1]['lat']},{paradas[-1]['lon']}"
    waypoints = "|".join(
        f"{p['lat']},{p['lon']}" for p in paradas[1:-1]
    )
    url = f"https://www.google.com/maps/dir/?api=1&origin={origin}&destination={dest}"
    if waypoints:
        url += f"&waypoints={urllib.parse.quote(waypoints)}"
    url += "&travelmode=driving"
    return url


def estimar_tempo_rota(dist_km, num_paradas, vel_media, tempo_parada):
    """Estima tempo total da rota (deslocamento + coletas)."""
    tempo_deslocamento = (dist_km / max(vel_media, 1)) * 60  # minutos
    tempo_coletas = num_paradas * tempo_parada
    return round(tempo_deslocamento + tempo_coletas)


def formatar_tempo(minutos):
    """Formata minutos em Xh Ymin."""
    if minutos < 60:
        return f"{minutos}min"
    h = int(minutos // 60)
    m = int(minutos % 60)
    return f"{h}h {m:02d}min"


def calcular_etas(paradas, hora_inicio_str, vel_media, tempo_parada):
    """Calcula horário estimado de chegada em cada parada."""
    try:
        h, m = map(int, hora_inicio_str.split(":"))
        t = datetime.now().replace(hour=h, minute=m, second=0, microsecond=0)
    except Exception:
        t = datetime.now().replace(hour=7, minute=30, second=0, microsecond=0)
    etas = []
    for i, p in enumerate(paradas):
        if i > 0:
            dist_km = p.get("dist_trecho_km", 0) or 0
            mins_traj = (dist_km / max(vel_media, 1)) * 60
            t += timedelta(minutes=mins_traj + tempo_parada)
        etas.append(t.strftime("%H:%M"))
    return etas


def gerar_link_whatsapp(motorista_nome, paradas, dist_km, hora_inicio, tel=""):
    """Gera link do WhatsApp com roteiro resumido."""
    linhas = [
        f"🚚 *ROTEIRO {motorista_nome.upper()}*",
        f"⏰ Saída: {hora_inicio}",
        f"📍 {max(len(paradas)-2, 0)} paradas | {dist_km:.1f} km",
        "",
    ]
    for i, p in enumerate(paradas):
        if i == 0:
            linhas.append(f"🏭 BASE: {p['nome']}")
        elif i == len(paradas) - 1:
            linhas.append(f"🔚 Retorno: {p['nome']}")
        else:
            linhas.append(f"{i}. {p['nome']}")
    texto = "\n".join(linhas)
    num = tel.strip().replace(" ", "").replace("-", "").replace("(", "").replace(")", "")
    if num and not num.startswith("+"):
        num = "55" + num
    base = f"https://wa.me/{num}" if num else "https://wa.me/"
    return f"{base}?text={urllib.parse.quote(texto)}"


def exportar_excel_rotas(rotas, hora_inicio, vel_media, tempo_parada):
    """Exporta rotas em Excel formatado com múltiplas sheets."""
    wb = Workbook()

    # ── Cores
    AZUL = "0a1628"
    AZUL_CLARO = "3498db"
    VERDE = "2ecc71"
    CINZA = "f0f2f6"
    BRANCO = "FFFFFF"

    def _header_style(cell, bg=AZUL, fg=BRANCO, bold=True, size=11):
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(color=fg, bold=bold, size=size)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _auto_width(ws, min_w=8, max_w=45):
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            w = max((len(str(c.value or "")) for c in col), default=min_w)
            ws.column_dimensions[col_letter].width = min(max(w + 2, min_w), max_w)

    # ─── Sheet Resumo ───────────────────────────────────────────────
    ws_res = wb.active
    ws_res.title = "Resumo"
    headers_res = ["Motorista", "Coletas", "Dist. (km)", "Litros Est.", "Custo R$", "Saída", "Retorno Est.", "Tempo Est."]
    for ci, h in enumerate(headers_res, 1):
        c = ws_res.cell(row=1, column=ci, value=h)
        _header_style(c)

    cores_mot = ["FFF3CD", "D4EFDF", "D6EAF8", "FAE5D3", "E8DAEF",
                 "D1F2EB", "FDEBD0", "FDEDEC", "EAF2FF", "F4ECF7"]
    for ri, r in enumerate(rotas, 2):
        etas = calcular_etas(r["paradas"], hora_inicio, vel_media, tempo_parada)
        retorno = etas[-1] if etas else "-"
        tempo_est = estimar_tempo_rota(r["distancia_km"], r["num_coletas"], vel_media, tempo_parada)
        vals = [r["motorista"], r["num_coletas"], r["distancia_km"],
                r.get("litros_estimados", 0), r.get("custo_combustivel", 0),
                hora_inicio, retorno, formatar_tempo(tempo_est)]
        for ci, v in enumerate(vals, 1):
            c = ws_res.cell(row=ri, column=ci, value=v)
            cor_bg = cores_mot[(ri - 2) % len(cores_mot)]
            c.fill = PatternFill("solid", fgColor=cor_bg)
            c.alignment = Alignment(horizontal="center", vertical="center")
    _auto_width(ws_res)

    # ─── Sheet por Motorista ─────────────────────────────────────────
    for v_idx, r in enumerate(rotas):
        etas = calcular_etas(r["paradas"], hora_inicio, vel_media, tempo_parada)
        sheet_name = r["motorista"][:31].replace("/", "-").replace(":", "")
        ws = wb.create_sheet(title=sheet_name)

        # Título
        ws.merge_cells("A1:I1")
        titulo = ws["A1"]
        titulo.value = f"ROTEIRO — {r['motorista'].upper()} | Saída: {hora_inicio} | {r['distancia_km']} km | {r['num_coletas']} coletas"
        _header_style(titulo, bg=AZUL_CLARO, size=12)
        ws.row_dimensions[1].height = 22

        headers = ["#", "Horário", "Local", "Endereço", "Obs", "Trecho (km)", "Acumulado (km)", "Lat", "Lon"]
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=2, column=ci, value=h)
            _header_style(c, bg=AZUL)

        alt_fill = PatternFill("solid", fgColor=CINZA)
        for pi, (p, eta) in enumerate(zip(r["paradas"], etas), 3):
            is_base = p["ordem"] == 0 or p["ordem"] == len(r["paradas"]) - 1
            vals = [
                p["ordem"], eta, p["nome"], p.get("endereco", ""),
                p.get("obs", ""),
                p.get("dist_trecho_km", "-") if p["ordem"] > 0 else "-",
                p.get("dist_acumulada_km", 0), p["lat"], p["lon"],
            ]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=pi, column=ci, value=v)
                c.alignment = Alignment(horizontal="center" if ci in (1, 2, 6, 7, 8, 9) else "left",
                                        vertical="center")
                if is_base:
                    c.fill = PatternFill("solid", fgColor="FFF9C4")
                    c.font = Font(bold=True)
                elif pi % 2 == 1:
                    c.fill = alt_fill
        _auto_width(ws)

    # ─── Sheet Mapa de Calor / Estatísticas ─────────────────────────
    ws_stat = wb.create_sheet(title="Estatísticas")
    ws_stat.merge_cells("A1:D1")
    t = ws_stat["A1"]
    t.value = "RESUMO ESTATÍSTICO — SKW COLETAS"
    _header_style(t, size=13)
    ws_stat.row_dimensions[1].height = 24

    stat_headers = ["Indicador", "Valor"]
    for ci, h in enumerate(stat_headers, 1):
        c = ws_stat.cell(row=2, column=ci, value=h)
        _header_style(c)

    dist_total = sum(r["distancia_km"] for r in rotas)
    col_total = sum(r["num_coletas"] for r in rotas)
    custo_total = sum(r.get("custo_combustivel", 0) for r in rotas)
    lit_total = sum(r.get("litros_estimados", 0) for r in rotas)
    stats = [
        ("Data/Hora Gerado", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Total de Motoristas", len(rotas)),
        ("Total de Coletas", col_total),
        ("Distância Total (km)", f"{dist_total:.2f}"),
        ("Litros Estimados", f"{lit_total:.2f} L"),
        ("Custo Total Combustível", f"R$ {custo_total:.2f}"),
        ("Custo por Coleta", f"R$ {custo_total/max(col_total,1):.2f}"),
        ("Distância Média/Motorista", f"{dist_total/max(len(rotas),1):.2f} km"),
        ("Coletas Médias/Motorista", f"{col_total/max(len(rotas),1):.1f}"),
    ]
    alt = PatternFill("solid", fgColor=CINZA)
    for ri, (ind, val) in enumerate(stats, 3):
        ws_stat.cell(row=ri, column=1, value=ind).alignment = Alignment(horizontal="left")
        c = ws_stat.cell(row=ri, column=2, value=val)
        c.alignment = Alignment(horizontal="center")
        if ri % 2 == 1:
            ws_stat.cell(row=ri, column=1).fill = alt
            c.fill = alt
    _auto_width(ws_stat)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def exportar_folha_impressao(rotas, hora_inicio, vel_media, tempo_parada):
    """Gera HTML de folha de rota para impressão."""
    html_parts = []
    html_parts.append("""<!DOCTYPE html><html><head><meta charset="utf-8">
<title>Roteiro SKW COLETAS</title>
<style>
body{font-family:Arial,sans-serif;margin:0;padding:16px;font-size:12px;}
h1{color:#0a1628;font-size:18px;margin:0 0 4px;}
.sub{color:#666;font-size:11px;margin-bottom:16px;}
.driver-block{page-break-inside:avoid;margin-bottom:24px;border:1px solid #ddd;border-radius:8px;overflow:hidden;}
.driver-header{background:#0a1628;color:#fff;padding:10px 14px;}
.driver-header h2{margin:0;font-size:14px;}
.driver-header p{margin:2px 0 0;font-size:11px;color:rgba(255,255,255,.7);}
table{width:100%;border-collapse:collapse;}
th{background:#3498db;color:#fff;padding:6px 8px;text-align:left;font-size:11px;}
td{padding:5px 8px;border-bottom:1px solid #eee;font-size:11px;}
tr:nth-child(even) td{background:#f8f9fb;}
tr.base-row td{background:#fffde7;font-weight:bold;}
.badge{display:inline-block;padding:2px 7px;border-radius:10px;font-size:10px;font-weight:bold;}
@media print{body{padding:0;}
.driver-block{page-break-inside:avoid;margin-bottom:12px;}
@page{margin:1cm;}}
</style></head><body>""")
    html_parts.append(f"<h1>📦 SKW COLETAS — Roteiro do Dia</h1>")
    html_parts.append(f'<div class="sub">Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")} &bull; Saída: {hora_inicio}</div>')

    for r in rotas:
        etas = calcular_etas(r["paradas"], hora_inicio, vel_media, tempo_parada)
        tempo_est = estimar_tempo_rota(r["distancia_km"], r["num_coletas"], vel_media, tempo_parada)
        html_parts.append(f"""<div class="driver-block">
<div class="driver-header">
  <h2>{r['motorista']}</h2>
  <p>{r['num_coletas']} coletas &bull; {r['distancia_km']} km &bull; {r.get('litros_estimados',0):.1f}L &bull; R${r.get('custo_combustivel',0):.2f} &bull; {formatar_tempo(tempo_est)}</p>
</div>
<table><tr><th>#</th><th>Horário</th><th>Local</th><th>Endereço</th><th>Obs</th><th>Trecho</th><th>Acumulado</th></tr>""")
        for p, eta in zip(r["paradas"], etas):
            is_base = p["ordem"] == 0 or p["ordem"] == len(r["paradas"]) - 1
            cls = ' class="base-row"' if is_base else ""
            trecho = f"{p.get('dist_trecho_km','0')} km" if p["ordem"] > 0 else "—"
            html_parts.append(f'<tr{cls}><td>{p["ordem"]}</td><td><b>{eta}</b></td>'
                               f'<td>{p["nome"]}</td><td>{p.get("endereco","")}</td>'
                               f'<td>{p.get("obs","")}</td><td>{trecho}</td>'
                               f'<td>{p.get("dist_acumulada_km",0):.2f} km</td></tr>')
        html_parts.append("</table></div>")

    html_parts.append('<script>window.onload=function(){window.print();}</script></body></html>')
    return "\n".join(html_parts)


# ═══════════════════════════════════════════════════════════════════════════
# CLUSTERIZACAO PROFISSIONAL
# ═══════════════════════════════════════════════════════════════════════════

def _dist_geo(lat1, lon1, lat2, lon2):
    return geodesic((lat1, lon1), (lat2, lon2)).meters


def _kmeans_geo(pontos, k, base_lat, base_lon, max_iter=80):
    n = len(pontos)
    coords = np.array([[p["lat"], p["lon"]] for p in pontos])
    angulos = [math.atan2(p["lat"] - base_lat, p["lon"] - base_lon) for p in pontos]
    idx_sorted = sorted(range(n), key=lambda i: angulos[i])
    step = max(n // k, 1)
    centroids = np.array([coords[idx_sorted[i * step % n]] for i in range(k)])
    labels = np.zeros(n, dtype=int)

    for _ in range(max_iter):
        for i in range(n):
            best_d = float("inf")
            for c in range(k):
                d = _dist_geo(coords[i][0], coords[i][1],
                              centroids[c][0], centroids[c][1])
                if d < best_d:
                    best_d = d
                    labels[i] = c
        new_centroids = np.copy(centroids)
        for c in range(k):
            mask = labels == c
            if mask.any():
                new_centroids[c] = coords[mask].mean(axis=0)
        if np.allclose(centroids, new_centroids, atol=1e-6):
            break
        centroids = new_centroids
    return labels


def agrupar_pontos_por_regiao(pontos, num_motoristas, base_lat, base_lon):
    if not pontos or num_motoristas < 1:
        return {}
    n = len(pontos)
    if num_motoristas >= n:
        return {i: [i] for i in range(n)}

    coords = np.array([[p["lat"], p["lon"]] for p in pontos])
    labels = _kmeans_geo(pontos, num_motoristas, base_lat, base_lon)

    # Passo 2: Outliers
    for _pass in range(5):
        moved = False
        for c in range(num_motoristas):
            idx_c = np.where(labels == c)[0]
            if len(idx_c) < 3:
                continue
            centroid = coords[idx_c].mean(axis=0)
            dists = np.array([_dist_geo(coords[i][0], coords[i][1],
                                        centroid[0], centroid[1]) for i in idx_c])
            threshold = dists.mean() * 1.8
            for j, i in enumerate(idx_c):
                if dists[j] <= threshold:
                    continue
                best_c, best_d = c, dists[j]
                for c2 in range(num_motoristas):
                    if c2 == c:
                        continue
                    idx_c2 = np.where(labels == c2)[0]
                    if len(idx_c2) == 0:
                        continue
                    cent2 = coords[idx_c2].mean(axis=0)
                    d2 = _dist_geo(coords[i][0], coords[i][1], cent2[0], cent2[1])
                    if d2 < best_d:
                        best_d = d2
                        best_c = c2
                if best_c != c:
                    labels[i] = best_c
                    moved = True
        if not moved:
            break

    # Passo 3: Fronteira
    for i in range(n):
        c_atual = labels[i]
        dists_to_clusters = []
        for c in range(num_motoristas):
            idx_c = np.where(labels == c)[0]
            if len(idx_c) == 0:
                dists_to_clusters.append(float("inf"))
                continue
            cent = coords[idx_c].mean(axis=0)
            dists_to_clusters.append(
                _dist_geo(coords[i][0], coords[i][1], cent[0], cent[1]))
        sorted_c = sorted(range(num_motoristas), key=lambda c: dists_to_clusters[c])
        c1, c2 = sorted_c[0], sorted_c[1]
        d1, d2 = dists_to_clusters[c1], dists_to_clusters[c2]
        if d1 > 0 and (d2 - d1) / d1 < 0.20 and c_atual != c1:
            def _raio_com(cluster_id, ponto_idx, incluir):
                idx = list(np.where(labels == cluster_id)[0])
                if incluir and ponto_idx not in idx:
                    idx.append(ponto_idx)
                elif not incluir and ponto_idx in idx:
                    idx.remove(ponto_idx)
                if len(idx) < 2:
                    return 0
                cent = coords[idx].mean(axis=0)
                return max(_dist_geo(coords[j][0], coords[j][1],
                                     cent[0], cent[1]) for j in idx)
            raio_c1_com = _raio_com(c1, i, True)
            raio_atual_sem = _raio_com(c_atual, i, False)
            raio_c1_sem = _raio_com(c1, i, False)
            raio_atual_com = _raio_com(c_atual, i, True)
            if raio_c1_com + raio_atual_sem < raio_c1_sem + raio_atual_com:
                labels[i] = c1

    # Passo 4: Balanceamento
    for _pass in range(n * 2):
        sizes = [int((labels == c).sum()) for c in range(num_motoristas)]
        max_c = max(range(num_motoristas), key=lambda c: sizes[c])
        min_c = min(range(num_motoristas), key=lambda c: sizes[c])
        if sizes[max_c] - sizes[min_c] <= 2:
            break
        idx_max = np.where(labels == max_c)[0]
        mask_min = labels == min_c
        cent_min = coords[mask_min].mean(axis=0) if mask_min.any() else coords.mean(axis=0)
        cent_max = coords[idx_max].mean(axis=0)
        dists_to_own = np.array([
            _dist_geo(coords[i][0], coords[i][1], cent_max[0], cent_max[1])
            for i in idx_max
        ])
        dists_to_target = np.array([
            _dist_geo(coords[i][0], coords[i][1], cent_min[0], cent_min[1])
            for i in idx_max
        ])
        scores = dists_to_target - dists_to_own * 0.5
        move_i = idx_max[np.argmin(scores)]
        labels[move_i] = min_c

    # Passo 5: Ordenar por angulo
    cluster_angles = []
    for c in range(num_motoristas):
        mask = labels == c
        if mask.any():
            cent = coords[mask].mean(axis=0)
            angle = math.atan2(cent[0] - base_lat, cent[1] - base_lon)
        else:
            angle = 999
        cluster_angles.append((c, angle))
    cluster_angles.sort(key=lambda x: x[1])
    remap = {old_c: new_c for new_c, (old_c, _) in enumerate(cluster_angles)}

    # Passo 6: Compacidade
    raios = {}
    for c in range(num_motoristas):
        idx_c = np.where(labels == c)[0]
        if len(idx_c) < 2:
            raios[c] = 0
            continue
        cent = coords[idx_c].mean(axis=0)
        raios[c] = max(_dist_geo(coords[i][0], coords[i][1],
                                  cent[0], cent[1]) for i in idx_c)
    raios_validos = [r for r in raios.values() if r > 0]
    if raios_validos:
        min_raio = min(raios_validos)
        for c in range(num_motoristas):
            if raios[c] > min_raio * 2.5:
                idx_c = np.where(labels == c)[0]
                if len(idx_c) <= 3:
                    continue
                cent = coords[idx_c].mean(axis=0)
                dists = [_dist_geo(coords[i][0], coords[i][1],
                                   cent[0], cent[1]) for i in idx_c]
                outlier_i = idx_c[np.argmax(dists)]
                best_c, best_d = c, max(dists)
                for c2 in range(num_motoristas):
                    if c2 == c:
                        continue
                    idx_c2 = np.where(labels == c2)[0]
                    if len(idx_c2) == 0:
                        continue
                    cent2 = coords[idx_c2].mean(axis=0)
                    d2 = _dist_geo(coords[outlier_i][0], coords[outlier_i][1],
                                   cent2[0], cent2[1])
                    if d2 < best_d:
                        best_d = d2
                        best_c = c2
                if best_c != c:
                    labels[outlier_i] = best_c

    grupos = {c: [] for c in range(num_motoristas)}
    for i in range(n):
        new_c = remap[labels[i]]
        grupos[new_c].append(i)
    return grupos


# ═══════════════════════════════════════════════════════════════════════════
# POS-PROCESSAMENTO: 2-OPT + OR-OPT
# ═══════════════════════════════════════════════════════════════════════════

def _two_opt(rota, matriz):
    n = len(rota)
    if n < 4:
        return rota
    melhorou = True
    while melhorou:
        melhorou = False
        for i in range(1, n - 2):
            for j in range(i + 1, n - 1):
                custo_atual = (matriz[rota[i-1]][rota[i]] +
                               matriz[rota[j]][rota[j+1] if j+1 < n else rota[0]])
                custo_novo = (matriz[rota[i-1]][rota[j]] +
                              matriz[rota[i]][rota[j+1] if j+1 < n else rota[0]])
                if custo_novo < custo_atual - 1:
                    rota[i:j+1] = rota[i:j+1][::-1]
                    melhorou = True
    return rota


def _or_opt(rota, matriz):
    n = len(rota)
    if n < 4:
        return rota
    for block_size in [1, 2, 3]:
        melhorou = True
        while melhorou:
            melhorou = False
            for i in range(1, n - block_size):
                bloco = rota[i:i + block_size]
                if i + block_size >= n:
                    continue
                antes = rota[i - 1]
                depois = rota[i + block_size] if i + block_size < n else rota[0]
                custo_remover = (
                    matriz[antes][bloco[0]] +
                    matriz[bloco[-1]][depois] -
                    matriz[antes][depois]
                )
                for j in range(1, n):
                    if abs(j - i) <= block_size:
                        continue
                    if j >= n:
                        continue
                    prev_j = rota[j - 1] if j > 0 else rota[-1]
                    next_j = rota[j] if j < n else rota[0]
                    custo_inserir = (
                        matriz[prev_j][bloco[0]] +
                        matriz[bloco[-1]][next_j] -
                        matriz[prev_j][next_j]
                    )
                    if custo_inserir < custo_remover - 10:
                        nova_rota = rota[:i] + rota[i + block_size:]
                        ins = j if j < i else j - block_size
                        ins = max(1, min(ins, len(nova_rota)))
                        nova_rota = nova_rota[:ins] + bloco + nova_rota[ins:]
                        rota = nova_rota
                        n = len(rota)
                        melhorou = True
                        break
                if melhorou:
                    break
    return rota


def criar_convex_hull(pontos_coords):
    if len(pontos_coords) < 3:
        return None
    try:
        points = MultiPoint([(lon, lat) for lat, lon in pontos_coords])
        hull = points.convex_hull
        if hull.geom_type == "Polygon":
            return hull
    except Exception:
        pass
    return None


# ═══════════════════════════════════════════════════════════════════════════
# OTIMIZACAO DE ROTA PROFISSIONAL
# ═══════════════════════════════════════════════════════════════════════════

def otimizar_rota(pontos, num_veiculos=1, deposito=0, max_dist_km=300,
                  balancear=True, preco_litro=6.29, km_por_litro=10.0,
                  tempo_busca_seg=15, capacidades=None, demandas=None):
    if len(pontos) < 2:
        return None

    matriz = calcular_matriz_distancias(pontos)
    dists_flat = matriz[matriz > 0]
    mediana_dist = float(np.median(dists_flat)) if len(dists_flat) > 0 else 5000
    limiar = mediana_dist * 1.2

    manager = pywrapcp.RoutingIndexManager(len(pontos), num_veiculos, deposito)
    routing = pywrapcp.RoutingModel(manager)

    def distance_callback(from_index, to_index):
        from_node = manager.IndexToNode(from_index)
        to_node = manager.IndexToNode(to_index)
        dist = int(matriz[from_node][to_node])
        if dist > limiar:
            excesso = (dist - limiar) / limiar
            penalidade = int(dist * excesso * excesso)
            return dist + penalidade
        return dist

    transit_callback_index = routing.RegisterTransitCallback(distance_callback)
    routing.SetArcCostEvaluatorOfAllVehicles(transit_callback_index)

    def distance_real_callback(from_index, to_index):
        return int(matriz[manager.IndexToNode(from_index)]
                         [manager.IndexToNode(to_index)])

    real_transit_idx = routing.RegisterTransitCallback(distance_real_callback)
    routing.AddDimension(
        real_transit_idx, 0, int(max_dist_km * 1000), True, "RealDistance",
    )
    if balancear and num_veiculos > 1:
        routing.GetDimensionOrDie("RealDistance").SetGlobalSpanCostCoefficient(150)

    # ── Capacidade de veículo (CVRP)
    if capacidades and demandas and any(c > 0 for c in capacidades):
        def demand_callback(from_index):
            node = manager.IndexToNode(from_index)
            return int(demandas[node]) if node < len(demandas) else 0
        demand_idx = routing.RegisterUnaryTransitCallback(demand_callback)
        routing.AddDimensionWithVehicleCapacity(
            demand_idx, 0,
            [int(c) if c and c > 0 else 999999 for c in capacidades],
            True, "Capacity",
        )

    best_solution = None
    best_cost = float("inf")
    strategies = [
        (routing_enums_pb2.FirstSolutionStrategy.PATH_CHEAPEST_ARC,
         routing_enums_pb2.LocalSearchMetaheuristic.GUIDED_LOCAL_SEARCH),
        (routing_enums_pb2.FirstSolutionStrategy.SAVINGS,
         routing_enums_pb2.LocalSearchMetaheuristic.GUIDED_LOCAL_SEARCH),
        (routing_enums_pb2.FirstSolutionStrategy.CHRISTOFIDES,
         routing_enums_pb2.LocalSearchMetaheuristic.SIMULATED_ANNEALING),
    ]
    time_per = max(tempo_busca_seg // len(strategies), 5)

    for first_strat, meta in strategies:
        sp = pywrapcp.DefaultRoutingSearchParameters()
        sp.first_solution_strategy = first_strat
        sp.local_search_metaheuristic = meta
        sp.time_limit.FromSeconds(time_per)
        sp.log_search = False
        solution = routing.SolveWithParameters(sp)
        if solution and solution.ObjectiveValue() < best_cost:
            best_cost = solution.ObjectiveValue()
            best_solution = solution

    solution = best_solution
    if not solution:
        return None

    rotas = []
    for vehicle_id in range(num_veiculos):
        rota_nodes = []
        index = routing.Start(vehicle_id)
        while not routing.IsEnd(index):
            rota_nodes.append(manager.IndexToNode(index))
            index = solution.Value(routing.NextVar(index))
        rota_nodes.append(manager.IndexToNode(index))

        pontos_internos = [n for n in rota_nodes if n != deposito]
        if not pontos_internos:
            continue

        if len(pontos_internos) >= 3:
            sub = [deposito] + pontos_internos + [deposito]
            sub = _two_opt(sub, matriz)
            sub = _or_opt(sub, matriz)
            rota_nodes = sub if sub[0] == deposito else [deposito] + sub
            if rota_nodes[-1] != deposito:
                rota_nodes.append(deposito)

        trechos = []
        distancia_total = 0
        for i in range(len(rota_nodes) - 1):
            node = rota_nodes[i]
            next_node = rota_nodes[i + 1]
            dist_trecho = matriz[node][next_node]
            distancia_total += dist_trecho
            trechos.append({
                "de_idx": node, "para_idx": next_node,
                "de_nome": pontos[node].get("nome", f"Ponto {node}"),
                "para_nome": pontos[next_node].get("nome", f"Ponto {next_node}"),
                "de_lat": pontos[node]["lat"], "de_lon": pontos[node]["lon"],
                "para_lat": pontos[next_node]["lat"], "para_lon": pontos[next_node]["lon"],
                "distancia_m": round(dist_trecho),
                "distancia_km": round(dist_trecho / 1000, 2),
            })

        paradas = []
        acumulado = 0.0
        for seq, node_idx in enumerate(rota_nodes):
            if seq > 0:
                acumulado += trechos[seq - 1]["distancia_km"]
            paradas.append({
                "ordem": seq, "node_idx": node_idx,
                "nome": pontos[node_idx].get("nome", f"Ponto {node_idx}"),
                "lat": pontos[node_idx]["lat"], "lon": pontos[node_idx]["lon"],
                "endereco": pontos[node_idx].get("endereco", ""),
                "obs": pontos[node_idx].get("obs", ""),
                "dist_trecho_km": trechos[seq - 1]["distancia_km"] if seq > 0 else 0.0,
                "dist_acumulada_km": round(acumulado, 2),
            })

        num_coletas = len([n for n in rota_nodes if n != deposito]) - (
            1 if rota_nodes[-1] == deposito else 0
        )
        dist_km = round(distancia_total / 1000, 2)
        litros = round(dist_km / km_por_litro, 2)
        custo = round(litros * preco_litro, 2)

        rotas.append({
            "rota": rota_nodes, "trechos": trechos, "paradas": paradas,
            "distancia_km": dist_km, "num_coletas": num_coletas,
            "litros_estimados": litros, "custo_combustivel": custo,
            "km_por_litro": km_por_litro, "preco_litro": preco_litro,
            "motorista": f"Motorista {vehicle_id + 1}",
        })
    return rotas


def geocodificar(endereco):
    geolocator = Nominatim(user_agent="skw_coletas_app")
    try:
        location = geolocator.geocode(f"{endereco}, Goiania, GO, Brasil")
        if location:
            return {"lat": location.latitude, "lon": location.longitude}
    except Exception:
        pass
    return None


def geocodificar_em_lote(enderecos, progresso_cb=None):
    """Geocodifica lista de endereços, retorna lista de pontos."""
    geolocator = Nominatim(user_agent="skw_coletas_lote")
    resultados = []
    for i, row in enumerate(enderecos):
        end = row.get("endereco", row.get("address", row.get("endereço", "")))
        nome = row.get("nome", row.get("name", row.get("local", end)))
        if progresso_cb:
            progresso_cb(i, len(enderecos), end)
        try:
            loc = geolocator.geocode(f"{end}, Goiania, GO, Brasil", timeout=5)
            if loc:
                resultados.append({
                    "lat": loc.latitude, "lon": loc.longitude,
                    "nome": str(nome), "endereco": str(end),
                    "obs": str(row.get("obs", row.get("observacao", ""))),
                    "prioridade": str(row.get("prioridade", "Normal")),
                    "_geocodificado": True,
                })
            else:
                resultados.append({"nome": str(nome), "endereco": str(end),
                                   "_geocodificado": False, "_erro": "Não encontrado"})
        except Exception as e:
            resultados.append({"nome": str(nome), "endereco": str(end),
                               "_geocodificado": False, "_erro": str(e)})
    return resultados


def salvar_sessao():
    """Serializa o estado da sessão para JSON."""
    estado = {
        "versao": "2.0",
        "data": datetime.now().isoformat(),
        "pontos_coleta": st.session_state.pontos_coleta,
        "motoristas": st.session_state.motoristas,
        "atribuicao_motorista": {str(k): v for k, v in st.session_state.atribuicao_motorista.items()},
        "bairros_selecionados": st.session_state.bairros_selecionados,
        "hora_inicio": st.session_state.hora_inicio,
        "status_pontos": {str(k): v for k, v in st.session_state.status_pontos.items()},
        "num_pacotes": {str(k): v for k, v in st.session_state.num_pacotes.items()},
        "capacidade_veiculos": {str(k): v for k, v in st.session_state.capacidade_veiculos.items()},
        "base_location": st.session_state.base_location,
        "tempo_coleta_min": st.session_state.tempo_coleta_min,
        "velocidade_media": st.session_state.velocidade_media,
    }
    return json.dumps(estado, ensure_ascii=False, indent=2)


def carregar_sessao(json_str):
    """Restaura o estado da sessão de um JSON."""
    try:
        estado = json.loads(json_str)
        st.session_state.pontos_coleta = estado.get("pontos_coleta", [])
        st.session_state.motoristas = estado.get("motoristas", [])
        raw_atrib = estado.get("atribuicao_motorista", {})
        st.session_state.atribuicao_motorista = {int(k): v for k, v in raw_atrib.items()}
        st.session_state.bairros_selecionados = estado.get("bairros_selecionados", {})
        st.session_state.hora_inicio = estado.get("hora_inicio", "07:30")
        raw_status = estado.get("status_pontos", {})
        st.session_state.status_pontos = {int(k): v for k, v in raw_status.items()}
        raw_pac = estado.get("num_pacotes", {})
        st.session_state.num_pacotes = {int(k): v for k, v in raw_pac.items()}
        raw_cap = estado.get("capacidade_veiculos", {})
        st.session_state.capacidade_veiculos = {int(k): v for k, v in raw_cap.items()}
        st.session_state.base_location = estado.get("base_location", GOIANIA_CENTER)
        st.session_state.tempo_coleta_min = estado.get("tempo_coleta_min", TEMPO_MEDIO_COLETA_MIN)
        st.session_state.velocidade_media = estado.get("velocidade_media", VELOCIDADE_MEDIA_KMH)
        st.session_state.rota_otimizada = None
        return True, f"Sessão carregada: {len(st.session_state.pontos_coleta)} pontos, {len(st.session_state.motoristas)} motoristas"
    except Exception as e:
        return False, f"Erro ao carregar sessão: {e}"


def carregar_geojson_bairros(arquivo):
    try:
        content = arquivo.read().decode("utf-8")
        return json.loads(content)
    except Exception as e:
        st.error(f"Erro ao carregar GeoJSON: {e}")
        return None


def ponto_no_poligono(lat, lon, geojson):
    ponto = Point(lon, lat)
    for feature in geojson.get("features", []):
        try:
            geom = shape(feature["geometry"])
            if geom.contains(ponto):
                nome = feature.get("properties", {}).get(
                    "name", feature.get("properties", {}).get(
                        "NOME", feature.get("properties", {}).get("nome", "Sem nome")))
                return nome
        except Exception:
            continue
    return "Fora dos limites"


def _get_bairro_nome(feature):
    """Extrai nome do bairro de um feature GeoJSON."""
    props = feature.get("properties", {})
    for key in ["name", "NOME", "nome", "NAME", "bairro", "BAIRRO"]:
        if key in props and props[key]:
            return str(props[key])
    return f"Bairro {id(feature) % 1000}"


def criar_mapa(pontos, rota=None, geojson_bairros=None, poligonos_custom=None,
               motoristas=None, atribuicao=None, bairros_sel=None):
    """Cria mapa com regioes coloridas por motorista e bairros selecionaveis."""
    m = folium.Map(location=GOIANIA_CENTER, zoom_start=12, tiles="CartoDB positron")

    folium.TileLayer("OpenStreetMap", name="OpenStreetMap").add_to(m)
    folium.TileLayer(
        "https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}",
        attr="Esri", name="Satelite",
    ).add_to(m)

    # ── Bairros com cores selecionadas
    if geojson_bairros:
        fg_bairros = folium.FeatureGroup(name="Bairros", show=True)
        for i, feature in enumerate(geojson_bairros.get("features", [])):
            nome = _get_bairro_nome(feature)

            if bairros_sel and nome in bairros_sel:
                cor_idx = bairros_sel[nome]
                cor = CORES_BAIRROS[cor_idx % len(CORES_BAIRROS)]
                opacity = 0.30
                weight = 3
            else:
                cor = "#95a5a6"
                opacity = 0.05
                weight = 1

            folium.GeoJson(
                feature, name=nome,
                style_function=lambda x, c=cor, o=opacity, w=weight: {
                    "fillColor": c, "color": c, "weight": w, "fillOpacity": o,
                },
                tooltip=nome,
            ).add_to(fg_bairros)
        fg_bairros.add_to(m)

    # ── Regioes dos motoristas (convex hull)
    if motoristas and atribuicao and len(pontos) > 1:
        fg_regioes = folium.FeatureGroup(name="Regioes Motoristas", show=True)
        pontos_por_mot = {}
        for p_idx, m_idx in atribuicao.items():
            if m_idx not in pontos_por_mot:
                pontos_por_mot[m_idx] = []
            real_idx = int(p_idx) + 1
            if real_idx < len(pontos):
                pontos_por_mot[m_idx].append(
                    (pontos[real_idx]["lat"], pontos[real_idx]["lon"]))

        for m_idx, coords in pontos_por_mot.items():
            if m_idx >= len(motoristas):
                continue
            cor = CORES_MOTORISTAS[m_idx % len(CORES_MOTORISTAS)]
            nome_mot = motoristas[m_idx]["nome"]
            if len(coords) >= 3:
                hull = criar_convex_hull(coords)
                if hull:
                    hull_expanded = hull.buffer(0.003)
                    coords_hull = list(hull_expanded.exterior.coords)
                    folium.Polygon(
                        locations=[[lat, lon] for lon, lat in coords_hull],
                        color=cor, weight=3, fill=True,
                        fill_color=cor, fill_opacity=0.15,
                        tooltip=f"Regiao: {nome_mot}",
                        popup=f"<b>{nome_mot}</b><br>{len(coords)} pontos",
                    ).add_to(fg_regioes)
            elif len(coords) == 2:
                folium.PolyLine(
                    locations=coords, color=cor, weight=4, opacity=0.6,
                    dash_array="10", tooltip=f"Regiao: {nome_mot}",
                ).add_to(fg_regioes)
        fg_regioes.add_to(m)

    # ── Poligonos customizados
    if poligonos_custom:
        for poly in poligonos_custom:
            folium.Polygon(
                locations=poly["coords"], color=poly.get("cor", "#ff7800"),
                weight=2, fill=True, fill_color=poly.get("cor", "#ff7800"),
                fill_opacity=0.2, tooltip=poly.get("nome", "Regiao"),
            ).add_to(m)

    # ── Base
    if pontos:
        folium.Marker(
            location=[pontos[0]["lat"], pontos[0]["lon"]],
            popup=f"<b>BASE SKW:</b> {pontos[0].get('nome', 'Deposito')}",
            icon=folium.Icon(color="red", icon="home", prefix="fa"),
            tooltip="Base SKW",
        ).add_to(m)

    # ── Pontos de coleta
    for i, p in enumerate(pontos[1:], 1):
        p_idx = i - 1
        mot_idx = atribuicao.get(p_idx, None) if atribuicao else None

        if mot_idx is not None and motoristas and mot_idx < len(motoristas):
            cor_folium = FOLIUM_COLORS[mot_idx % len(FOLIUM_COLORS)]
            cor_hex = CORES_MOTORISTAS[mot_idx % len(CORES_MOTORISTAS)]
            nome_mot = motoristas[mot_idx]["nome"]
            motorista_info = f"<br><span style='color:{cor_hex};font-weight:bold'>■ {nome_mot}</span>"
        else:
            cor_folium = "gray"
            motorista_info = "<br><i>Sem motorista</i>"

        prio = p.get("prioridade", "Normal")
        prio_html = ""
        if prio == "Alta":
            prio_html = "<br><span style='color:#e74c3c;font-weight:bold'>⚡ PRIORIDADE ALTA</span>"
        elif prio == "Urgente":
            prio_html = "<br><span style='color:#c0392b;font-weight:bold'>🔥 URGENTE</span>"

        popup_html = f"""
        <div style="min-width:220px;font-family:sans-serif">
            <b style="font-size:14px">#{i} - {p.get('nome', f'Ponto {i}')}</b>
            {motorista_info}{prio_html}
            <br><small>Lat: {p['lat']:.6f} | Lon: {p['lon']:.6f}</small>
            {f"<br>End: {p.get('endereco', '')}" if p.get('endereco') else ""}
            {f"<br>Obs: {p.get('obs', '')}" if p.get('obs') else ""}
            <br><a href='https://www.google.com/maps?q={p["lat"]},{p["lon"]}' target='_blank'>
            Abrir no Google Maps</a>
        </div>
        """
        folium.Marker(
            location=[p["lat"], p["lon"]],
            popup=folium.Popup(popup_html, max_width=320),
            icon=folium.Icon(color=cor_folium, icon="box", prefix="fa"),
            tooltip=f"#{i} - {p.get('nome', f'Ponto {i}')}",
        ).add_to(m)

    # ── Rota otimizada
    if rota:
        for v_idx, r in enumerate(rota):
            cor = CORES_MOTORISTAS[v_idx % len(CORES_MOTORISTAS)]
            coords_rota = []
            for node_idx in r["rota"]:
                coords_rota.append([pontos[node_idx]["lat"], pontos[node_idx]["lon"]])

            folium.PolyLine(
                coords_rota, color=cor, weight=5, opacity=0.85,
                tooltip=f"{r.get('motorista', f'Veiculo {v_idx+1}')} - {r['distancia_km']} km",
            ).add_to(m)

            for seq, node_idx in enumerate(r["rota"]):
                if seq == 0:
                    continue
                folium.CircleMarker(
                    location=[pontos[node_idx]["lat"], pontos[node_idx]["lon"]],
                    radius=14, color=cor, fill=True, fill_color=cor,
                    fill_opacity=0.9, tooltip=f"Parada {seq}",
                ).add_to(m)
                folium.Marker(
                    location=[pontos[node_idx]["lat"], pontos[node_idx]["lon"]],
                    icon=folium.DivIcon(
                        html=f'<div style="font-size:11px;color:white;font-weight:bold;'
                             f'text-align:center;line-height:28px">{seq}</div>',
                        icon_size=(28, 28), icon_anchor=(14, 14),
                    ),
                ).add_to(m)

    Draw(
        draw_options={
            "polyline": False, "rectangle": True, "polygon": True,
            "circle": False, "marker": True, "circlemarker": False,
        },
        edit_options={"edit": True, "remove": True},
    ).add_to(m)

    folium.LayerControl(collapsed=False).add_to(m)
    return m


# ══════════════════════════════════════════════════════════════════════════════
# CSS SKW COLETAS
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<style>
.skw-header {
    background: linear-gradient(135deg, #0a1628 0%, #1a2744 40%, #0d3b66 100%);
    padding: 30px 36px;
    border-radius: 16px;
    margin-bottom: 20px;
    box-shadow: 0 10px 40px rgba(0,0,0,.22);
    position: relative;
    overflow: hidden;
}
.skw-header::before {
    content: "";
    position: absolute;
    top: -60%; right: -5%;
    width: 350px; height: 350px;
    background: radial-gradient(circle, rgba(52,152,219,.12) 0%, transparent 70%);
    border-radius: 50%;
}
.skw-header::after {
    content: "";
    position: absolute;
    bottom: -40%; left: 10%;
    width: 200px; height: 200px;
    background: radial-gradient(circle, rgba(46,204,113,.08) 0%, transparent 70%);
    border-radius: 50%;
}
.skw-header .skw-brand {
    display: flex;
    align-items: center;
    gap: 14px;
}
.skw-header .skw-logo {
    font-size: 2.2rem;
    background: rgba(255,255,255,.1);
    padding: 10px 14px;
    border-radius: 12px;
}
.skw-header h1 {
    color: #fff;
    margin: 0;
    font-size: 2rem;
    font-weight: 800;
    letter-spacing: 1px;
}
.skw-header .skw-sub {
    color: rgba(255,255,255,.5);
    font-size: .85rem;
    margin-top: 2px;
    letter-spacing: .5px;
}
.skw-header .skw-stats {
    display: flex;
    gap: 28px;
    margin-top: 14px;
}
.skw-header .skw-stat {
    color: rgba(255,255,255,.75);
    font-size: .88rem;
}
.skw-header .skw-stat b {
    color: #fff;
    font-size: 1.1rem;
}

.kpi-row { display: flex; gap: 12px; flex-wrap: wrap; margin: 14px 0; }
.kpi-card {
    flex: 1 1 150px;
    background: #fff;
    border: 1px solid #e8ecf1;
    border-radius: 12px;
    padding: 16px 18px;
    text-align: center;
    box-shadow: 0 2px 8px rgba(0,0,0,.04);
    transition: transform .15s, box-shadow .15s;
}
.kpi-card:hover { transform: translateY(-2px); box-shadow: 0 6px 20px rgba(0,0,0,.08); }
.kpi-card .kpi-icon { font-size: 1.5rem; margin-bottom: 2px; }
.kpi-card .kpi-value { font-size: 1.4rem; font-weight: 700; color: #0a1628; line-height: 1.2; }
.kpi-card .kpi-label { font-size: .72rem; color: #7f8c8d; text-transform: uppercase; letter-spacing: .5px; margin-top: 2px; }
.kpi-card.green  { border-top: 3px solid #2ecc71; }
.kpi-card.blue   { border-top: 3px solid #3498db; }
.kpi-card.orange { border-top: 3px solid #f39c12; }
.kpi-card.red    { border-top: 3px solid #e74c3c; }
.kpi-card.purple { border-top: 3px solid #9b59b6; }
.kpi-card.teal   { border-top: 3px solid #1abc9c; }

.mot-card {
    display: flex; align-items: center; gap: 14px;
    padding: 12px 16px; background: #fff;
    border: 1px solid #e8ecf1; border-radius: 10px;
    margin-bottom: 6px; transition: box-shadow .15s;
}
.mot-card:hover { box-shadow: 0 4px 12px rgba(0,0,0,.07); }
.mot-badge {
    width: 40px; height: 40px; border-radius: 50%;
    display: flex; align-items: center; justify-content: center;
    color: #fff; font-weight: 700; font-size: 1rem; flex-shrink: 0;
}
.mot-info { flex: 1; }
.mot-info .mot-name { font-weight: 600; font-size: .95rem; color: #0a1628; }
.mot-info .mot-detail { font-size: .78rem; color: #7f8c8d; }
.mot-pontos {
    background: #f0f2f6; padding: 3px 10px; border-radius: 20px;
    font-size: .78rem; font-weight: 600; color: #2c3e50;
}

.section-title {
    font-size: 1.1rem; font-weight: 700; color: #0a1628;
    margin: 20px 0 8px 0; padding-bottom: 6px;
    border-bottom: 2px solid #3498db; display: inline-block;
}
.section-subtitle { font-size: .85rem; color: #7f8c8d; margin: -4px 0 12px 0; }

.map-legend { display: flex; flex-wrap: wrap; gap: 6px; padding: 8px 0; }
.map-legend-item {
    display: flex; align-items: center; gap: 5px;
    padding: 4px 10px; background: #f8f9fb;
    border: 1px solid #e8ecf1; border-radius: 20px;
    font-size: .78rem; font-weight: 500;
}
.map-legend-dot { width: 10px; height: 10px; border-radius: 50%; flex-shrink: 0; }

.status-bar {
    display: flex; gap: 18px; padding: 8px 16px;
    background: #f8f9fb; border: 1px solid #e8ecf1;
    border-radius: 8px; margin: 6px 0 12px 0;
    font-size: .82rem; color: #2c3e50;
}
.status-bar b { color: #0a1628; }

.eco-banner {
    background: linear-gradient(135deg, #d4efdf 0%, #a9dfbf 100%);
    border: 1px solid #82e0aa; border-radius: 12px;
    padding: 14px 20px; margin: 10px 0;
}
.eco-banner .eco-title { font-weight: 700; color: #1e8449; font-size: .95rem; margin-bottom: 4px; }
.eco-banner .eco-values { display: flex; gap: 20px; flex-wrap: wrap; }
.eco-banner .eco-item { font-size: .85rem; color: #1a5032; }
.eco-banner .eco-item b { font-size: 1rem; }

.rota-header-card {
    display: flex; align-items: center; gap: 12px;
    padding: 12px 16px; background: #fff;
    border: 1px solid #e8ecf1; border-left: 4px solid;
    border-radius: 10px; margin-bottom: 6px;
}
.rota-header-card .rota-icon { font-size: 1.6rem; }
.rota-header-card .rota-info { flex: 1; }
.rota-header-card .rota-name { font-weight: 700; font-size: 1rem; color: #0a1628; }
.rota-header-card .rota-stats { font-size: .78rem; color: #7f8c8d; margin-top: 2px; }

.trajeto-line {
    display: flex; flex-wrap: wrap; align-items: center;
    gap: 3px; padding: 8px 0; font-size: .78rem;
}
.trajeto-stop {
    background: #eaf2f8; padding: 2px 8px; border-radius: 5px;
    font-weight: 600; color: #2c3e50; white-space: nowrap;
}
.trajeto-arrow { color: #bdc3c7; font-size: .65rem; }
.trajeto-dist { color: #7f8c8d; font-size: .68rem; font-style: italic; }

.bairro-chip {
    display: inline-flex; align-items: center; gap: 5px;
    padding: 4px 12px; border-radius: 20px;
    font-size: .82rem; font-weight: 500;
    margin: 2px; cursor: default;
    border: 2px solid;
}

.footer-pro {
    text-align: center; padding: 16px; margin-top: 24px;
    border-top: 1px solid #e8ecf1; color: #95a5a6; font-size: .75rem;
}

/* ── Agenda / Timeline ────────────────────────────────────────── */
.agenda-card {
    background: #fff; border: 1px solid #e8ecf1;
    border-radius: 12px; padding: 16px 20px; margin-bottom: 14px;
    box-shadow: 0 2px 8px rgba(0,0,0,.04);
}
.agenda-card .ag-driver {
    display: flex; align-items: center; gap: 10px; margin-bottom: 12px;
}
.agenda-card .ag-badge {
    width: 36px; height: 36px; border-radius: 50%;
    display: flex; align-items: center; justify-content: center;
    color: #fff; font-weight: 700; font-size: .9rem; flex-shrink: 0;
}
.agenda-card .ag-name { font-weight: 700; font-size: 1rem; color: #0a1628; }
.agenda-card .ag-sub { font-size: .78rem; color: #7f8c8d; }
.timeline {
    position: relative; padding-left: 28px;
    border-left: 3px solid #e8ecf1; margin-left: 14px;
}
.tl-item {
    position: relative; margin-bottom: 10px; padding: 8px 12px;
    background: #f8f9fb; border-radius: 8px;
}
.tl-item::before {
    content: ""; position: absolute; left: -19px; top: 50%;
    transform: translateY(-50%);
    width: 12px; height: 12px; border-radius: 50%;
    background: #3498db; border: 2px solid #fff;
    box-shadow: 0 0 0 2px #3498db;
}
.tl-item.base-item::before { background: #e74c3c; box-shadow: 0 0 0 2px #e74c3c; }
.tl-item .tl-time {
    font-weight: 700; font-size: .88rem; color: #0a1628; display: block;
}
.tl-item .tl-name { font-size: .82rem; color: #2c3e50; margin-top: 1px; }
.tl-item .tl-dist { font-size: .72rem; color: #95a5a6; }

/* ── Status Badges ─────────────────────────────────────────────── */
.status-badge {
    display: inline-flex; align-items: center; gap: 4px;
    padding: 3px 10px; border-radius: 12px;
    font-size: .75rem; font-weight: 600; white-space: nowrap;
}
.status-pendente  { background: #fef9e7; color: #d68910; border: 1px solid #f9ca24; }
.status-coletado  { background: #eafaf1; color: #1e8449; border: 1px solid #2ecc71; }
.status-falhou    { background: #fdedec; color: #922b21; border: 1px solid #e74c3c; }
.status-reagendado{ background: #eaf2ff; color: #1a5276; border: 1px solid #3498db; }

/* ── Progress bar de status ────────────────────────────────────── */
.prog-bar-outer {
    background: #e8ecf1; border-radius: 6px; height: 10px;
    overflow: hidden; margin: 4px 0;
}
.prog-bar-inner {
    height: 100%; border-radius: 6px; transition: width .3s;
    background: linear-gradient(90deg, #2ecc71, #1abc9c);
}

/* ── Capacidade chip ────────────────────────────────────────────── */
.cap-chip {
    display: inline-flex; align-items: center; gap: 4px;
    padding: 2px 8px; border-radius: 8px; font-size: .72rem; font-weight: 600;
    background: #eaf2ff; color: #1a5276; border: 1px solid #85c1e9;
}
.cap-chip.over { background: #fdedec; color: #922b21; border-color: #e74c3c; }

section[data-testid="stSidebar"] { background: #f8f9fb; }
section[data-testid="stSidebar"] .stMarkdown h2 {
    color: #0a1628; font-size: 1rem;
    border-bottom: 2px solid #3498db; padding-bottom: 4px;
}
button[data-baseweb="tab"] { font-weight: 600 !important; font-size: .85rem !important; }
#MainMenu { visibility: hidden; }
footer { visibility: hidden; }
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# HEADER SKW COLETAS
# ══════════════════════════════════════════════════════════════════════════════
n_pts = len(st.session_state.pontos_coleta)
n_mot = len(st.session_state.motoristas)
n_poly = len(st.session_state.poligonos)
n_bairros = len(st.session_state.bairros_selecionados)
tem_rota = st.session_state.rota_otimizada is not None

st.markdown(f"""
<div class="skw-header">
    <div class="skw-brand">
        <div class="skw-logo">📦</div>
        <div>
            <h1>SKW COLETAS</h1>
            <div class="skw-sub">SISTEMA DE ROTEIRIZACAO E COLETAS &bull; GOIANIA/GO</div>
        </div>
    </div>
    <div class="skw-stats">
        <div class="skw-stat"><b>{n_pts}</b> pontos</div>
        <div class="skw-stat"><b>{n_mot}</b> motoristas</div>
        <div class="skw-stat"><b>{n_bairros}</b> bairros</div>
        {"<div class='skw-stat'>✅ <b>Rota otimizada</b></div>" if tem_rota else ""}
    </div>
</div>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("## SKW COLETAS")

    with st.expander("Base (Ponto de Partida)", expanded=True):
        base_endereco = st.text_input("Endereco da base", value="Centro, Goiania, GO")
        sb1, sb2 = st.columns(2)
        with sb1:
            base_lat = st.number_input("Latitude", value=-16.6869, format="%.6f", key="blat")
        with sb2:
            base_lon = st.number_input("Longitude", value=-49.2648, format="%.6f", key="blon")
        if st.button("Geocodificar Base", use_container_width=True):
            loc = geocodificar(base_endereco)
            if loc:
                st.session_state.base_location = [loc["lat"], loc["lon"]]
                st.success(f"Base: {loc['lat']:.6f}, {loc['lon']:.6f}")
            else:
                st.error("Endereco nao encontrado")

    with st.expander("Importar Pontos", expanded=not bool(n_pts)):
        arquivo_pontos = st.file_uploader(
            "CSV, XLSX ou JSON", type=["csv", "xlsx", "xls", "json"],
            key="upload_pontos", label_visibility="collapsed",
        )
        if arquivo_pontos:
            try:
                if arquivo_pontos.name.endswith(".csv"):
                    df = pd.read_csv(arquivo_pontos)
                elif arquivo_pontos.name.endswith((".xlsx", ".xls")):
                    df = pd.read_excel(arquivo_pontos)
                else:
                    df = pd.DataFrame(json.loads(arquivo_pontos.read().decode("utf-8")))
                col_map = {}
                for col in df.columns:
                    cl = col.lower().strip()
                    if cl in ("latitude", "lat"):
                        col_map[col] = "lat"
                    elif cl in ("longitude", "lon", "lng", "long"):
                        col_map[col] = "lon"
                    elif cl in ("nome", "name", "local"):
                        col_map[col] = "nome"
                    elif cl in ("endereco", "endereço", "address"):
                        col_map[col] = "endereco"
                    elif cl in ("obs", "observacao", "observação", "nota"):
                        col_map[col] = "obs"
                    elif cl in ("prioridade", "priority"):
                        col_map[col] = "prioridade"
                df = df.rename(columns=col_map)
                if "lat" in df.columns and "lon" in df.columns:
                    pontos_importados = []
                    for _, row in df.iterrows():
                        pontos_importados.append({
                            "lat": float(row["lat"]),
                            "lon": float(row["lon"]),
                            "nome": str(row.get("nome", "")),
                            "endereco": str(row.get("endereco", "")),
                            "obs": str(row.get("obs", "")),
                            "prioridade": str(row.get("prioridade", "Normal")),
                        })
                    st.success(f"{len(pontos_importados)} pontos importados!")
                    st.session_state.pontos_coleta = pontos_importados
                    st.session_state.atribuicao_motorista = {}
                else:
                    st.error("Colunas 'lat' e 'lon' nao encontradas.")
            except Exception as e:
                st.error(f"Erro: {e}")

    with st.expander("Bairros (GeoJSON)"):
        arquivo_geojson = st.file_uploader(
            "Arquivo GeoJSON", type=["geojson", "json"], key="upload_geojson",
            label_visibility="collapsed",
        )
        geojson_bairros = st.session_state.geojson_bairros_data
        if arquivo_geojson:
            geojson_bairros = carregar_geojson_bairros(arquivo_geojson)
            if geojson_bairros:
                st.session_state.geojson_bairros_data = geojson_bairros
                st.success(f"{len(geojson_bairros.get('features', []))} poligonos!")

        if st.button("Baixar Bairros (OSM)", use_container_width=True):
            with st.spinner("Baixando..."):
                try:
                    import osmnx as ox
                    gdf = ox.features_from_place(
                        "Goiania, Goias, Brazil",
                        tags={"admin_level": "10", "boundary": "administrative"},
                    )
                    if len(gdf) == 0:
                        gdf = ox.features_from_place(
                            "Goiania, Goias, Brazil",
                            tags={"place": ["neighbourhood", "suburb"]},
                        )
                    if len(gdf) > 0:
                        gdf_poly = gdf[gdf.geometry.type.isin(["Polygon", "MultiPolygon"])]
                        if len(gdf_poly) > 0:
                            geojson_str = gdf_poly.to_json()
                            geojson_bairros = json.loads(geojson_str)
                            st.session_state.geojson_bairros_data = geojson_bairros
                            st.success(f"{len(gdf_poly)} bairros!")
                            st.download_button("Salvar GeoJSON", geojson_str,
                                               "bairros_goiania.geojson", "application/json")
                except Exception as e:
                    st.error(f"Erro: {e}")

    with st.expander("⏰ Horário e Agenda"):
        st.session_state.hora_inicio = st.text_input(
            "Hora de saída (HH:MM)", value=st.session_state.hora_inicio,
            key="hora_inicio_sb", placeholder="07:30",
        )
        st.caption("Usado para calcular horários estimados em cada parada.")

    with st.expander("💾 Salvar / Carregar Sessão"):
        st.caption("Salve todo o trabalho (pontos, motoristas, atribuições) em um arquivo JSON.")
        if st.button("💾 Baixar Sessão Atual", use_container_width=True):
            sessao_json = salvar_sessao()
            st.download_button(
                "⬇️ Clique para baixar",
                sessao_json,
                f"skw_sessao_{datetime.now().strftime('%Y%m%d_%H%M')}.json",
                "application/json",
                use_container_width=True,
            )
        arq_sessao = st.file_uploader(
            "Carregar sessão (.json)", type=["json"], key="upload_sessao",
            label_visibility="collapsed",
        )
        if arq_sessao:
            ok, msg = carregar_sessao(arq_sessao.read().decode("utf-8"))
            if ok:
                st.success(msg)
                st.rerun()
            else:
                st.error(msg)

    with st.expander("🌐 Geocodificação em Lote"):
        st.caption("Envie CSV com coluna 'endereco' (e opcionalmente 'nome', 'obs', 'prioridade').")
        arq_geocod = st.file_uploader(
            "CSV de endereços", type=["csv"], key="upload_geocod",
            label_visibility="collapsed",
        )
        if arq_geocod and st.button("Geocodificar Todos", use_container_width=True, type="primary"):
            try:
                df_gc = pd.read_csv(arq_geocod)
                df_gc.columns = [c.lower().strip() for c in df_gc.columns]
                registros = df_gc.to_dict("records")
                prog_bar = st.progress(0, text="Geocodificando…")

                def _prog(i, total, end):
                    prog_bar.progress((i + 1) / max(total, 1), text=f"Geocodificando: {end[:40]}…")

                resultados = geocodificar_em_lote(registros, _prog)
                prog_bar.empty()
                ok_list = [r for r in resultados if r.get("_geocodificado")]
                fail_list = [r for r in resultados if not r.get("_geocodificado")]

                if ok_list:
                    for p in ok_list:
                        p.pop("_geocodificado", None)
                        st.session_state.pontos_coleta.append(p)
                    st.success(f"✅ {len(ok_list)} pontos geocodificados e adicionados!")
                if fail_list:
                    st.warning(f"⚠️ {len(fail_list)} endereços não encontrados: "
                               + ", ".join(r.get("endereco","?") for r in fail_list[:5]))
                st.rerun()
            except Exception as e:
                st.error(f"Erro: {e}")

    with st.expander("Combustivel e Tempos"):
        preco_litro = st.number_input(
            "Preco (R$/L)", min_value=1.0, max_value=15.0,
            value=6.29, step=0.10, format="%.2f", key="preco_litro_sb",
        )
        km_por_litro = st.number_input(
            "Consumo (km/L)", min_value=3.0, max_value=30.0,
            value=10.0, step=0.5, format="%.1f", key="km_litro_sb",
        )
        max_dist = st.number_input(
            "Dist. maxima/veiculo (km)",
            min_value=10, max_value=1000, value=300, step=10, key="max_dist_sb",
        )
        tempo_busca = st.select_slider(
            "Nivel de otimizacao",
            options=[10, 20, 30, 60], value=20,
            format_func=lambda x: {10: "Rapido", 20: "Normal", 30: "Profundo", 60: "Maximo"}[x],
            key="tempo_sb",
        )
        st.divider()
        st.session_state.tempo_coleta_min = st.number_input(
            "Tempo por coleta (min)", min_value=1, max_value=60,
            value=st.session_state.tempo_coleta_min, key="tempo_col_sb",
        )
        st.session_state.velocidade_media = st.number_input(
            "Velocidade media (km/h)", min_value=10, max_value=80,
            value=st.session_state.velocidade_media, key="vel_sb",
        )


# ══════════════════════════════════════════════════════════════════════════════
# DASHBOARD KPI
# ══════════════════════════════════════════════════════════════════════════════
rotas_ativas = st.session_state.rota_otimizada
if rotas_ativas:
    _dist_t = sum(r["distancia_km"] for r in rotas_ativas)
    _lit_t = sum(r.get("litros_estimados", 0) for r in rotas_ativas)
    _cust_t = sum(r.get("custo_combustivel", 0) for r in rotas_ativas)
    _col_t = sum(r["num_coletas"] for r in rotas_ativas)
    _tempo_t = sum(estimar_tempo_rota(
        r["distancia_km"], r["num_coletas"],
        st.session_state.velocidade_media, st.session_state.tempo_coleta_min
    ) for r in rotas_ativas)
    st.markdown(f"""
    <div class="kpi-row">
        <div class="kpi-card blue"><div class="kpi-icon">📍</div>
            <div class="kpi-value">{n_pts}</div><div class="kpi-label">Pontos</div></div>
        <div class="kpi-card green"><div class="kpi-icon">👷</div>
            <div class="kpi-value">{len(rotas_ativas)}</div><div class="kpi-label">Motoristas</div></div>
        <div class="kpi-card orange"><div class="kpi-icon">🛣️</div>
            <div class="kpi-value">{_dist_t:.1f} km</div><div class="kpi-label">Distancia</div></div>
        <div class="kpi-card red"><div class="kpi-icon">⛽</div>
            <div class="kpi-value">{_lit_t:.1f} L</div><div class="kpi-label">Combustivel</div></div>
        <div class="kpi-card purple"><div class="kpi-icon">💰</div>
            <div class="kpi-value">R$ {_cust_t:.2f}</div><div class="kpi-label">Custo</div></div>
        <div class="kpi-card teal"><div class="kpi-icon">⏱️</div>
            <div class="kpi-value">{formatar_tempo(_tempo_t)}</div><div class="kpi-label">Tempo Est.</div></div>
    </div>
    """, unsafe_allow_html=True)
else:
    st.markdown(f"""
    <div class="kpi-row">
        <div class="kpi-card blue"><div class="kpi-icon">📍</div>
            <div class="kpi-value">{n_pts}</div><div class="kpi-label">Pontos</div></div>
        <div class="kpi-card green"><div class="kpi-icon">👷</div>
            <div class="kpi-value">{n_mot}</div><div class="kpi-label">Motoristas</div></div>
        <div class="kpi-card orange"><div class="kpi-icon">🏘️</div>
            <div class="kpi-value">{n_bairros}</div><div class="kpi-label">Bairros</div></div>
    </div>
    """, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# ABAS
# ══════════════════════════════════════════════════════════════════════════════
tab_mapa, tab_rota, tab_motoristas, tab_bairros, tab_pontos, tab_agenda, tab_status, tab_relatorio = st.tabs([
    "🗺️ Mapa", "🚀 Roteirização", "👷 Motoristas",
    "🏘️ Bairros", "📍 Pontos", "📅 Agenda", "📊 Status", "📋 Relatório",
])


# ══════════════════════════════════════════════════════════════════════════════
# TAB MAPA
# ══════════════════════════════════════════════════════════════════════════════
with tab_mapa:
    todos_pontos = [
        {"lat": base_lat, "lon": base_lon, "nome": "BASE SKW"}
    ] + st.session_state.pontos_coleta

    mapa = criar_mapa(
        todos_pontos,
        rota=st.session_state.rota_otimizada,
        geojson_bairros=geojson_bairros,
        poligonos_custom=st.session_state.poligonos,
        motoristas=st.session_state.motoristas,
        atribuicao=st.session_state.atribuicao_motorista,
        bairros_sel=st.session_state.bairros_selecionados,
    )

    map_data = st_folium(mapa, width=None, height=620, key="main_map")

    if map_data and map_data.get("all_drawings"):
        drawings = map_data["all_drawings"]
        new_polys = []
        new_points = []
        for d in drawings:
            geom = d.get("geometry", {})
            if geom.get("type") == "Polygon":
                coords = geom["coordinates"][0]
                new_polys.append({
                    "coords": [[c[1], c[0]] for c in coords],
                    "cor": "#ff7800",
                    "nome": f"Regiao desenhada {len(new_polys)+1}",
                })
            elif geom.get("type") == "Point":
                lon, lat = geom["coordinates"]
                new_points.append({"lat": lat, "lon": lon, "nome": "Ponto manual"})
        if new_polys:
            st.session_state.poligonos = new_polys
        if new_points:
            for p in new_points:
                if p not in st.session_state.pontos_coleta:
                    st.session_state.pontos_coleta.append(p)

    if st.session_state.motoristas:
        leg_items = ""
        for i, mot in enumerate(st.session_state.motoristas):
            cor = mot["cor"]
            pontos_mot = sum(1 for v in st.session_state.atribuicao_motorista.values() if v == i)
            leg_items += (
                f'<div class="map-legend-item">'
                f'<div class="map-legend-dot" style="background:{cor}"></div>'
                f'<span>{mot["nome"]} ({pontos_mot})</span></div>'
            )
        st.markdown(f'<div class="map-legend">{leg_items}</div>', unsafe_allow_html=True)

    st.markdown(
        f'<div class="status-bar">'
        f'<span>📍 <b>{n_pts}</b> pontos</span>'
        f'<span>👷 <b>{n_mot}</b> motoristas</span>'
        f'<span>🏘️ <b>{n_bairros}</b> bairros</span>'
        f'{"<span>✅ <b>Rota otimizada</b></span>" if tem_rota else ""}'
        f'</div>', unsafe_allow_html=True,
    )


# ══════════════════════════════════════════════════════════════════════════════
# TAB ROTEIRIZACAO
# ══════════════════════════════════════════════════════════════════════════════
with tab_rota:
    st.markdown('<div class="section-title">Roteirizacao SKW</div>', unsafe_allow_html=True)
    st.markdown('<div class="section-subtitle">OR-Tools + 2-opt + Or-opt | Penalidade cubica | Multi-estrategia</div>', unsafe_allow_html=True)

    if len(st.session_state.pontos_coleta) < 2:
        st.warning("Adicione pelo menos 2 pontos para otimizar.")
    elif not st.session_state.motoristas:
        st.warning("Cadastre motoristas na aba 'Motoristas' primeiro.")
    else:
        num_mot = len(st.session_state.motoristas)

        col_modo, col_bal = st.columns([3, 1])
        with col_modo:
            modo_rota = st.radio(
                "Modo", ["Por motorista (usa atribuicao)", "Otimizacao global (OR-Tools divide)"],
                horizontal=True,
            )
        with col_bal:
            balancear = st.checkbox("Balancear carga", value=True)

        if st.button("OTIMIZAR ROTAS", type="primary", use_container_width=True):
            with st.spinner("Calculando rotas otimizadas... Aguarde."):
                if modo_rota == "Por motorista (usa atribuicao)":
                    if not st.session_state.atribuicao_motorista:
                        n_m = len(st.session_state.motoristas)
                        pontos = st.session_state.pontos_coleta
                        grupos = agrupar_pontos_por_regiao(pontos, n_m, base_lat, base_lon)
                        nova_atrib = {}
                        for m_idx, indices in grupos.items():
                            for p_idx in indices:
                                nova_atrib[p_idx] = m_idx
                        st.session_state.atribuicao_motorista = nova_atrib

                    if st.session_state.atribuicao_motorista:
                        todas_rotas = []
                        base = {"lat": base_lat, "lon": base_lon, "nome": "BASE SKW"}
                        for m_idx, mot in enumerate(st.session_state.motoristas):
                            pontos_idx = [k for k, v in st.session_state.atribuicao_motorista.items() if v == m_idx]
                            if not pontos_idx:
                                continue
                            pontos_mot = [base] + [st.session_state.pontos_coleta[i] for i in pontos_idx]
                            resultado = otimizar_rota(
                                pontos_mot, 1, deposito=0, max_dist_km=max_dist,
                                balancear=False, preco_litro=preco_litro,
                                km_por_litro=km_por_litro, tempo_busca_seg=tempo_busca,
                            )
                            if resultado:
                                r = resultado[0]
                                r["motorista"] = mot["nome"]
                                r["motorista_idx"] = m_idx
                                r["cor"] = mot["cor"]
                                todas_rotas.append(r)

                        if todas_rotas:
                            todos_pontos_global = [base] + st.session_state.pontos_coleta
                            rotas_globais = []
                            for rota_mot in todas_rotas:
                                m_idx = rota_mot["motorista_idx"]
                                pontos_idx = sorted([k for k, v in st.session_state.atribuicao_motorista.items() if v == m_idx])
                                mapa_local_global = {0: 0}
                                for local_i, global_i in enumerate(pontos_idx, 1):
                                    mapa_local_global[local_i] = global_i + 1
                                nova_rota = [mapa_local_global.get(nd, nd) for nd in rota_mot["rota"]]
                                novas_paradas = []
                                for p in rota_mot["paradas"]:
                                    p_copy = dict(p)
                                    p_copy["node_idx"] = mapa_local_global.get(p["node_idx"], p["node_idx"])
                                    novas_paradas.append(p_copy)
                                rotas_globais.append({**rota_mot, "rota": nova_rota, "paradas": novas_paradas})
                            st.session_state.rota_otimizada = rotas_globais
                            st.session_state.todos_pontos_rota = todos_pontos_global
                            st.rerun()
                        else:
                            st.error("Nenhuma rota encontrada.")
                else:
                    todos_pontos = [{"lat": base_lat, "lon": base_lon, "nome": "BASE SKW"}] + st.session_state.pontos_coleta
                    resultado = otimizar_rota(
                        todos_pontos, num_mot, deposito=0, max_dist_km=max_dist,
                        balancear=balancear, preco_litro=preco_litro,
                        km_por_litro=km_por_litro, tempo_busca_seg=tempo_busca,
                    )
                    if resultado:
                        for i, r in enumerate(resultado):
                            if i < len(st.session_state.motoristas):
                                r["motorista"] = st.session_state.motoristas[i]["nome"]
                                r["motorista_idx"] = i
                                r["cor"] = st.session_state.motoristas[i]["cor"]
                        st.session_state.rota_otimizada = resultado
                        st.session_state.todos_pontos_rota = todos_pontos
                        nova_atribuicao = {}
                        for r_idx, r in enumerate(resultado):
                            m_idx = r.get("motorista_idx", r_idx)
                            for node in r["rota"]:
                                if node > 0:
                                    nova_atribuicao[node - 1] = m_idx
                        st.session_state.atribuicao_motorista = nova_atribuicao
                        st.rerun()
                    else:
                        st.error("Sem solucao. Aumente distancia maxima ou motoristas.")

        # ── Resultados
        if st.session_state.rota_otimizada:
            rotas = st.session_state.rota_otimizada
            todos_pontos = st.session_state.get("todos_pontos_rota", [])
            dist_total = sum(r["distancia_km"] for r in rotas)
            coletas_total = sum(r["num_coletas"] for r in rotas)
            litros_total = sum(r.get("litros_estimados", 0) for r in rotas)
            custo_total = sum(r.get("custo_combustivel", 0) for r in rotas)

            # Economia
            if todos_pontos and len(todos_pontos) > 2:
                dist_seq = sum(geodesic(
                    (todos_pontos[i]["lat"], todos_pontos[i]["lon"]),
                    (todos_pontos[i+1]["lat"], todos_pontos[i+1]["lon"]),
                ).km for i in range(len(todos_pontos) - 1))
                dist_seq += geodesic(
                    (todos_pontos[-1]["lat"], todos_pontos[-1]["lon"]),
                    (todos_pontos[0]["lat"], todos_pontos[0]["lon"]),
                ).km
                if dist_seq > dist_total:
                    eco_km = dist_seq - dist_total
                    eco_pct = (eco_km / dist_seq) * 100
                    eco_rs = (eco_km / km_por_litro) * preco_litro
                    st.markdown(f"""
                    <div class="eco-banner">
                        <div class="eco-title">Economia com Otimizacao SKW</div>
                        <div class="eco-values">
                            <div class="eco-item"><b>{eco_km:.1f} km</b> a menos ({eco_pct:.0f}%)</div>
                            <div class="eco-item"><b>{eco_km / km_por_litro:.1f} L</b> economizados</div>
                            <div class="eco-item"><b>R$ {eco_rs:.2f}</b> de economia</div>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)

            # Tabela comparativa
            st.markdown('<div class="section-title">Divisao por Motorista</div>', unsafe_allow_html=True)
            df_comp = pd.DataFrame([{
                "Motorista": r["motorista"],
                "Coletas": r["num_coletas"],
                "Dist (km)": r["distancia_km"],
                "Litros": r.get("litros_estimados", 0),
                "Custo (R$)": r.get("custo_combustivel", 0),
                "Tempo Est.": formatar_tempo(estimar_tempo_rota(
                    r["distancia_km"], r["num_coletas"],
                    st.session_state.velocidade_media, st.session_state.tempo_coleta_min)),
                "% Dist": round(r["distancia_km"] / max(dist_total, 0.01) * 100, 1),
            } for r in rotas])
            st.dataframe(df_comp, use_container_width=True, hide_index=True,
                         column_config={
                             "Dist (km)": st.column_config.NumberColumn(format="%.2f"),
                             "Litros": st.column_config.NumberColumn(format="%.1f"),
                             "Custo (R$)": st.column_config.NumberColumn(format="R$ %.2f"),
                             "% Dist": st.column_config.ProgressColumn(min_value=0, max_value=100, format="%.1f%%"),
                         })

            # Roteiro detalhado
            st.markdown('<div class="section-title">Roteiro Detalhado</div>', unsafe_allow_html=True)
            for v_idx, r in enumerate(rotas):
                m_idx = r.get("motorista_idx", v_idx)
                icone = ICONES_MOTORISTAS[m_idx % len(ICONES_MOTORISTAS)]
                cor = CORES_MOTORISTAS[m_idx % len(CORES_MOTORISTAS)]
                maior_trecho = max((t["distancia_km"] for t in r["trechos"]), default=0) if r["trechos"] else 0
                tempo_est = estimar_tempo_rota(
                    r["distancia_km"], r["num_coletas"],
                    st.session_state.velocidade_media, st.session_state.tempo_coleta_min)

                st.markdown(f"""
                <div class="rota-header-card" style="border-left-color: {cor}">
                    <div class="rota-icon">{icone}</div>
                    <div class="rota-info">
                        <div class="rota-name">{r['motorista']}</div>
                        <div class="rota-stats">
                            {r['num_coletas']} coletas &bull; {r['distancia_km']} km &bull;
                            {r.get('litros_estimados',0)} L &bull; R$ {r.get('custo_combustivel',0):.2f} &bull;
                            ⏱️ {formatar_tempo(tempo_est)}
                        </div>
                    </div>
                    <div class="mot-pontos">{r['num_coletas']} pts</div>
                </div>
                """, unsafe_allow_html=True)

                with st.expander(f"Detalhes - {r['motorista']}", expanded=(v_idx == 0)):
                    # Trajeto visual
                    traj_html = '<div class="trajeto-line">'
                    for p_i, p in enumerate(r["paradas"]):
                        nome_curto = p["nome"][:22]
                        if p_i > 0:
                            traj_html += f'<span class="trajeto-dist">{r["trechos"][p_i-1]["distancia_km"]}km</span>'
                            traj_html += '<span class="trajeto-arrow">&#9654;</span>'
                        traj_html += f'<span class="trajeto-stop">{nome_curto}</span>'
                    traj_html += '</div>'
                    st.markdown(traj_html, unsafe_allow_html=True)

                    # Links ação
                    hora_ini = st.session_state.hora_inicio
                    vel_m = st.session_state.velocidade_media
                    t_col = st.session_state.tempo_coleta_min
                    etas = calcular_etas(r["paradas"], hora_ini, vel_m, t_col)

                    gmaps_link = gerar_link_google_maps(r["paradas"])
                    mot_tel = ""
                    m_idx_r = r.get("motorista_idx", v_idx)
                    if m_idx_r < len(st.session_state.motoristas):
                        mot_tel = st.session_state.motoristas[m_idx_r].get("telefone", "")
                    wa_link = gerar_link_whatsapp(r["motorista"], r["paradas"], r["distancia_km"], hora_ini, mot_tel)

                    lnk1, lnk2 = st.columns(2)
                    with lnk1:
                        if gmaps_link:
                            st.markdown(f"[🗺️ **Google Maps**]({gmaps_link})", unsafe_allow_html=False)
                    with lnk2:
                        st.markdown(f"[💬 **WhatsApp**]({wa_link})", unsafe_allow_html=False)

                    # Tabela com ETAs
                    df_paradas = pd.DataFrame([{
                        "#": p["ordem"],
                        "⏰ Horário": etas[pi] if pi < len(etas) else "-",
                        "Local": p["nome"],
                        "Endereco": p["endereco"],
                        "Trecho (km)": p["dist_trecho_km"] if p["ordem"] > 0 else "-",
                        "Acumulado (km)": p["dist_acumulada_km"],
                    } for pi, p in enumerate(r["paradas"])])
                    st.dataframe(df_paradas, use_container_width=True, hide_index=True)

            # Exportacao
            st.markdown('<div class="section-title">Exportar</div>', unsafe_allow_html=True)
            hora_exp = st.session_state.hora_inicio
            vel_exp = st.session_state.velocidade_media
            t_exp = st.session_state.tempo_coleta_min
            exp1, exp2, exp3, exp4, exp5 = st.columns(5)
            with exp1:
                all_rows = []
                for r in rotas:
                    etas_exp = calcular_etas(r["paradas"], hora_exp, vel_exp, t_exp)
                    for pi, p in enumerate(r["paradas"]):
                        all_rows.append({
                            "Motorista": r["motorista"], "Ordem": p["ordem"],
                            "Horario": etas_exp[pi] if pi < len(etas_exp) else "-",
                            "Local": p["nome"], "Endereco": p["endereco"],
                            "Obs": p["obs"], "Lat": p["lat"], "Lon": p["lon"],
                            "Trecho (km)": p["dist_trecho_km"],
                            "Acumulado (km)": p["dist_acumulada_km"],
                        })
                st.download_button("📄 CSV", pd.DataFrame(all_rows).to_csv(index=False),
                                   "skw_roteiro.csv", "text/csv", use_container_width=True)
            with exp2:
                try:
                    xlsx_bytes = exportar_excel_rotas(rotas, hora_exp, vel_exp, t_exp)
                    st.download_button(
                        "📊 Excel",
                        xlsx_bytes,
                        f"skw_roteiro_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
                except Exception as e:
                    st.error(f"Excel: {e}")
            with exp3:
                zip_buf = io.BytesIO()
                with zipfile.ZipFile(zip_buf, "w") as zf:
                    for r in rotas:
                        etas_z = calcular_etas(r["paradas"], hora_exp, vel_exp, t_exp)
                        rows = [{"Ordem": p["ordem"],
                                 "Horario": etas_z[pi] if pi < len(etas_z) else "-",
                                 "Local": p["nome"], "Endereco": p["endereco"],
                                 "Lat": p["lat"], "Lon": p["lon"],
                                 "Trecho (km)": p["dist_trecho_km"],
                                 "Acumulado (km)": p["dist_acumulada_km"]}
                                for pi, p in enumerate(r["paradas"])]
                        fname = r["motorista"].replace(" ", "_").lower()
                        zf.writestr(f"{fname}.csv", pd.DataFrame(rows).to_csv(index=False))
                st.download_button("🗜️ ZIP", zip_buf.getvalue(),
                                   "skw_motoristas.zip", "application/zip", use_container_width=True)
            with exp4:
                export_json = json.dumps([{
                    "motorista": r["motorista"], "distancia_km": r["distancia_km"],
                    "num_coletas": r["num_coletas"], "paradas": r["paradas"],
                } for r in rotas], ensure_ascii=False, indent=2)
                st.download_button("🔧 JSON", export_json,
                                   "skw_roteiro.json", "application/json", use_container_width=True)
            with exp5:
                html_imp = exportar_folha_impressao(rotas, hora_exp, vel_exp, t_exp)
                st.download_button(
                    "🖨️ Imprimir",
                    html_imp.encode("utf-8"),
                    "skw_roteiro_impressao.html",
                    "text/html",
                    use_container_width=True,
                    help="Baixe e abra no navegador para imprimir",
                )

            if st.button("Limpar Rotas", type="secondary"):
                st.session_state.rota_otimizada = None
                st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# TAB MOTORISTAS
# ══════════════════════════════════════════════════════════════════════════════
with tab_motoristas:
    st.markdown('<div class="section-title">Equipe de Motoristas</div>', unsafe_allow_html=True)
    st.markdown('<div class="section-subtitle">Cadastre ate 10 motoristas com cores exclusivas</div>', unsafe_allow_html=True)

    with st.expander("Adicionar Novo Motorista", expanded=not bool(st.session_state.motoristas)):
        fm1, fm2, fm3 = st.columns([2, 2, 1])
        with fm1:
            novo_mot_nome = st.text_input("Nome", key="novo_mot_nome", placeholder="Ex: Joao Silva")
        with fm2:
            novo_mot_tel = st.text_input("Telefone", key="novo_mot_tel", placeholder="(62) 99999-0000")
        with fm3:
            novo_mot_placa = st.text_input("Placa", key="novo_mot_placa", placeholder="ABC-1234")
        fmc1, fmc2 = st.columns([1, 1])
        with fmc1:
            novo_mot_cap = st.number_input(
                "Capacidade (volumes, 0=ilimitado)", min_value=0, max_value=9999,
                value=0, step=1, key="novo_mot_cap",
            )
        with fmc2:
            novo_mot_tipo = st.selectbox(
                "Tipo de veículo", ["Moto", "Carro", "Van", "Caminhão", "Outro"],
                key="novo_mot_tipo",
            )
        if st.button("Adicionar Motorista", type="primary",
                     disabled=len(st.session_state.motoristas) >= 10, use_container_width=True):
            if novo_mot_nome.strip():
                idx = len(st.session_state.motoristas)
                st.session_state.motoristas.append({
                    "nome": novo_mot_nome.strip(), "telefone": novo_mot_tel.strip(),
                    "placa": novo_mot_placa.strip(),
                    "cor": CORES_MOTORISTAS[idx % len(CORES_MOTORISTAS)],
                    "cor_nome": NOMES_CORES[idx % len(NOMES_CORES)],
                    "capacidade": int(novo_mot_cap),
                    "tipo_veiculo": novo_mot_tipo,
                })
                if novo_mot_cap > 0:
                    st.session_state.capacidade_veiculos[idx] = int(novo_mot_cap)
                st.rerun()
            else:
                st.error("Nome obrigatorio")

    if not st.session_state.motoristas:
        st.info("Nenhum motorista cadastrado.")

    remover_idx = None
    for i in range(10):
        if i < len(st.session_state.motoristas):
            mot = st.session_state.motoristas[i]
            pontos_atrib = sum(1 for v in st.session_state.atribuicao_motorista.values() if v == i)
            cap = mot.get("capacidade", 0)
            tipo = mot.get("tipo_veiculo", "")
            cap_html = f'<span class="cap-chip">📦 {cap} vol.</span>' if cap > 0 else '<span class="cap-chip">📦 Ilimitado</span>'
            tipo_icons = {"Moto": "🏍️", "Carro": "🚗", "Van": "🚐", "Caminhão": "🚛", "Outro": "🚚"}
            tipo_html = f' | {tipo_icons.get(tipo,"🚚")} {tipo}' if tipo else ""
            st.markdown(f"""
            <div class="mot-card">
                <div class="mot-badge" style="background:{mot['cor']}">{i+1}</div>
                <div class="mot-info">
                    <div class="mot-name">{mot['nome']}</div>
                    <div class="mot-detail">{mot['cor_nome']}{tipo_html}
                        {(' | ' + mot['telefone']) if mot.get('telefone') else ''}
                        {(' | ' + mot['placa']) if mot.get('placa') else ''}</div>
                </div>
                {cap_html}
                <div class="mot-pontos">{pontos_atrib} pts</div>
            </div>
            """, unsafe_allow_html=True)
            if st.button(f"Remover", key=f"rm_mot_{i}"):
                remover_idx = i

    if remover_idx is not None:
        st.session_state.motoristas.pop(remover_idx)
        new_atrib = {}
        for pk, mv in st.session_state.atribuicao_motorista.items():
            if mv == remover_idx:
                continue
            elif mv > remover_idx:
                new_atrib[pk] = mv - 1
            else:
                new_atrib[pk] = mv
        st.session_state.atribuicao_motorista = new_atrib
        st.rerun()

    # Atribuicao
    if st.session_state.motoristas and st.session_state.pontos_coleta:
        st.markdown('<div class="section-title">Atribuicao de Pontos</div>', unsafe_allow_html=True)
        atrib_mode = st.radio("Modo", ["Automatica (por regiao)", "Manual (ponto a ponto)"],
                              horizontal=True, label_visibility="collapsed")

        if atrib_mode == "Automatica (por regiao)":
            st.caption("Distribui pontos em regioes geograficas. Cada motorista recebe uma fatia da cidade.")
            if st.button("Distribuir Automaticamente", type="primary", use_container_width=True):
                n_m = len(st.session_state.motoristas)
                grupos = agrupar_pontos_por_regiao(st.session_state.pontos_coleta, n_m, base_lat, base_lon)
                nova_atribuicao = {}
                for m_idx, indices in grupos.items():
                    for p_idx in indices:
                        nova_atribuicao[p_idx] = m_idx
                st.session_state.atribuicao_motorista = nova_atribuicao
                st.rerun()
        else:
            nomes_mot = ["-- Sem motorista --"] + [
                f"{ICONES_MOTORISTAS[i]} {m['nome']}" for i, m in enumerate(st.session_state.motoristas)
            ]
            nova_atribuicao = dict(st.session_state.atribuicao_motorista)
            for p_idx, p in enumerate(st.session_state.pontos_coleta):
                current = st.session_state.atribuicao_motorista.get(p_idx, None)
                default_idx = (current + 1) if current is not None else 0
                sel = st.selectbox(
                    f"#{p_idx+1} - {p.get('nome', 'Ponto')[:40]}",
                    options=range(len(nomes_mot)), format_func=lambda x: nomes_mot[x],
                    index=default_idx, key=f"atrib_{p_idx}",
                )
                if sel == 0:
                    nova_atribuicao.pop(p_idx, None)
                else:
                    nova_atribuicao[p_idx] = sel - 1
            if st.button("Salvar Atribuicoes", use_container_width=True):
                st.session_state.atribuicao_motorista = nova_atribuicao
                st.rerun()

    if st.session_state.atribuicao_motorista and st.session_state.motoristas:
        st.markdown('<div class="section-title">Resumo</div>', unsafe_allow_html=True)
        resumo_data = []
        for m_idx, mot in enumerate(st.session_state.motoristas):
            pontos_idx = [k for k, v in st.session_state.atribuicao_motorista.items() if v == m_idx]
            resumo_data.append({"Cor": ICONES_MOTORISTAS[m_idx], "Motorista": mot["nome"], "Pontos": len(pontos_idx)})
        st.dataframe(pd.DataFrame(resumo_data), use_container_width=True, hide_index=True)


# ══════════════════════════════════════════════════════════════════════════════
# TAB BAIRROS
# ══════════════════════════════════════════════════════════════════════════════
with tab_bairros:
    st.markdown('<div class="section-title">Selecao de Bairros</div>', unsafe_allow_html=True)
    st.markdown('<div class="section-subtitle">Selecione bairros e atribua cores diferentes para visualizar no mapa</div>', unsafe_allow_html=True)

    if geojson_bairros and geojson_bairros.get("features"):
        nomes_bairros = sorted(set(
            _get_bairro_nome(f) for f in geojson_bairros["features"]
        ))

        # Cores disponiveis
        st.caption(f"{len(nomes_bairros)} bairros disponiveis | {len(st.session_state.bairros_selecionados)} selecionados")

        # Legenda de cores
        leg = ""
        for ci, cor in enumerate(CORES_BAIRROS):
            leg += f'<span class="bairro-chip" style="background:{cor}22;border-color:{cor};color:{cor}">{ci+1}</span>'
        st.markdown(f"**Paleta de cores:** {leg}", unsafe_allow_html=True)

        st.divider()

        # Busca
        busca = st.text_input("Buscar bairro", placeholder="Digite o nome do bairro...", key="busca_bairro")

        bairros_filtrados = [b for b in nomes_bairros if busca.lower() in b.lower()] if busca else nomes_bairros

        # Acoes em massa
        ac1, ac2, ac3 = st.columns(3)
        with ac1:
            if st.button("Selecionar Todos", use_container_width=True):
                for i, b in enumerate(nomes_bairros):
                    if b not in st.session_state.bairros_selecionados:
                        st.session_state.bairros_selecionados[b] = i % len(CORES_BAIRROS)
                st.rerun()
        with ac2:
            if st.button("Limpar Selecao", use_container_width=True):
                st.session_state.bairros_selecionados = {}
                st.rerun()
        with ac3:
            if st.button("Cores Automaticas", use_container_width=True, help="Atribui cores sequenciais"):
                for i, b in enumerate(st.session_state.bairros_selecionados.keys()):
                    st.session_state.bairros_selecionados[b] = i % len(CORES_BAIRROS)
                st.rerun()

        st.divider()

        # Lista de bairros com checkbox e cor
        nova_sel = dict(st.session_state.bairros_selecionados)
        for b_nome in bairros_filtrados:
            is_sel = b_nome in st.session_state.bairros_selecionados
            bc1, bc2 = st.columns([3, 1])
            with bc1:
                checked = st.checkbox(
                    b_nome, value=is_sel, key=f"bairro_cb_{b_nome}",
                )
            with bc2:
                if checked:
                    cor_idx = st.selectbox(
                        "Cor", options=list(range(len(CORES_BAIRROS))),
                        index=nova_sel.get(b_nome, 0) % len(CORES_BAIRROS),
                        format_func=lambda x: f"Cor {x+1}",
                        key=f"bairro_cor_{b_nome}",
                        label_visibility="collapsed",
                    )
                    nova_sel[b_nome] = cor_idx
                else:
                    nova_sel.pop(b_nome, None)

        if st.button("Aplicar Selecao de Bairros", type="primary", use_container_width=True):
            st.session_state.bairros_selecionados = nova_sel
            st.rerun()

        # Bairros selecionados
        if st.session_state.bairros_selecionados:
            st.markdown('<div class="section-title">Bairros Selecionados</div>', unsafe_allow_html=True)
            chips = ""
            for b_nome, cor_idx in st.session_state.bairros_selecionados.items():
                cor = CORES_BAIRROS[cor_idx % len(CORES_BAIRROS)]
                chips += f'<span class="bairro-chip" style="background:{cor}22;border-color:{cor};color:{cor}">{b_nome}</span>'
            st.markdown(chips, unsafe_allow_html=True)

            # Pontos por bairro
            if st.session_state.pontos_coleta:
                st.markdown('<div class="section-title">Pontos por Bairro Selecionado</div>', unsafe_allow_html=True)
                bairro_contagem = {}
                for p in st.session_state.pontos_coleta:
                    b = ponto_no_poligono(p["lat"], p["lon"], geojson_bairros)
                    if b in st.session_state.bairros_selecionados:
                        bairro_contagem[b] = bairro_contagem.get(b, 0) + 1
                if bairro_contagem:
                    df_bc = pd.DataFrame([{"Bairro": k, "Pontos": v} for k, v in sorted(bairro_contagem.items(), key=lambda x: -x[1])])
                    st.dataframe(df_bc, use_container_width=True, hide_index=True)
                else:
                    st.info("Nenhum ponto nos bairros selecionados.")
    else:
        st.info("Importe um arquivo GeoJSON de bairros na barra lateral para usar esta funcao.")


# ══════════════════════════════════════════════════════════════════════════════
# TAB PONTOS DE COLETA
# ══════════════════════════════════════════════════════════════════════════════
with tab_pontos:
    st.markdown('<div class="section-title">Pontos de Coleta</div>', unsafe_allow_html=True)

    with st.expander("Adicionar Ponto Manualmente"):
        p1, p2, p3 = st.columns(3)
        with p1:
            novo_nome = st.text_input("Nome", key="novo_nome", placeholder="Nome do ponto")
        with p2:
            novo_lat = st.number_input("Latitude", value=-16.6869, format="%.6f", key="nlat")
        with p3:
            novo_lon = st.number_input("Longitude", value=-49.2648, format="%.6f", key="nlon")
        p4, p5, p6 = st.columns([2, 2, 1])
        with p4:
            novo_end = st.text_input("Endereco", key="novo_end", placeholder="Opcional")
        with p5:
            nova_obs = st.text_input("Observacao", key="nova_obs", placeholder="Opcional")
        with p6:
            nova_prio = st.selectbox("Prioridade", ["Normal", "Alta", "Urgente"], key="nova_prio")

        btn1, btn2 = st.columns(2)
        with btn1:
            if st.button("Adicionar Ponto", use_container_width=True):
                st.session_state.pontos_coleta.append({
                    "lat": novo_lat, "lon": novo_lon,
                    "nome": novo_nome or f"Ponto {len(st.session_state.pontos_coleta)+1}",
                    "endereco": novo_end, "obs": nova_obs, "prioridade": nova_prio,
                })
                st.rerun()
        with btn2:
            if st.button("Geocodificar e Adicionar", use_container_width=True) and novo_end:
                loc = geocodificar(novo_end)
                if loc:
                    st.session_state.pontos_coleta.append({
                        "lat": loc["lat"], "lon": loc["lon"],
                        "nome": novo_nome or novo_end,
                        "endereco": novo_end, "obs": nova_obs, "prioridade": nova_prio,
                    })
                    st.rerun()
                else:
                    st.error("Endereco nao encontrado")

    if st.session_state.pontos_coleta:
        st.markdown(
            f'<div class="status-bar">'
            f'<span>📍 <b>{len(st.session_state.pontos_coleta)}</b> pontos</span>'
            f'<span>✅ <b>{len(st.session_state.atribuicao_motorista)}</b> atribuidos</span>'
            f'</div>', unsafe_allow_html=True,
        )
        pontos_display = []
        for i, p in enumerate(st.session_state.pontos_coleta):
            mot_idx = st.session_state.atribuicao_motorista.get(i, None)
            mot_nome = ""
            if mot_idx is not None and mot_idx < len(st.session_state.motoristas):
                mot_nome = st.session_state.motoristas[mot_idx]["nome"]
            pontos_display.append({**p, "motorista": mot_nome})
        df_pontos = pd.DataFrame(pontos_display)
        edited_df = st.data_editor(df_pontos, num_rows="dynamic", use_container_width=True, key="editor_pontos")

        c1, c2, c3, c4 = st.columns(4)
        with c1:
            if st.button("Salvar Alteracoes", use_container_width=True):
                records = edited_df.to_dict("records")
                st.session_state.pontos_coleta = [{k: v for k, v in r.items() if k != "motorista"} for r in records]
                st.rerun()
        with c2:
            st.download_button("CSV", df_pontos.to_csv(index=False),
                               "skw_pontos.csv", "text/csv", use_container_width=True)
        with c3:
            st.download_button("JSON", df_pontos.to_json(orient="records", force_ascii=False),
                               "skw_pontos.json", "application/json", use_container_width=True)
        with c4:
            if st.button("Limpar Pontos", type="secondary", use_container_width=True):
                st.session_state.pontos_coleta = []
                st.session_state.rota_otimizada = None
                st.session_state.atribuicao_motorista = {}
                st.rerun()
    else:
        st.info("Importe pontos pela barra lateral ou adicione manualmente.")


# ══════════════════════════════════════════════════════════════════════════════
# TAB AGENDA
# ══════════════════════════════════════════════════════════════════════════════
with tab_agenda:
    st.markdown('<div class="section-title">Agenda do Dia</div>', unsafe_allow_html=True)
    hora_ag = st.session_state.hora_inicio
    vel_ag = st.session_state.velocidade_media
    t_ag = st.session_state.tempo_coleta_min

    if not st.session_state.rota_otimizada:
        st.info("Otimize as rotas na aba Roteirização para visualizar a agenda.")
    else:
        rotas_ag = st.session_state.rota_otimizada
        st.markdown(f"**Saída:** {hora_ag} &nbsp;|&nbsp; **Vel. média:** {vel_ag} km/h &nbsp;|&nbsp; **Tempo por coleta:** {t_ag} min")
        st.divider()

        # Visão por motorista: timeline
        for v_idx, r in enumerate(rotas_ag):
            m_idx = r.get("motorista_idx", v_idx)
            cor = CORES_MOTORISTAS[m_idx % len(CORES_MOTORISTAS)]
            icone = ICONES_MOTORISTAS[m_idx % len(ICONES_MOTORISTAS)]
            etas_ag = calcular_etas(r["paradas"], hora_ag, vel_ag, t_ag)
            retorno = etas_ag[-1] if etas_ag else "-"
            tempo_est = estimar_tempo_rota(r["distancia_km"], r["num_coletas"], vel_ag, t_ag)

            st.markdown(f"""
            <div class="agenda-card">
              <div class="ag-driver">
                <div class="ag-badge" style="background:{cor}">{icone}</div>
                <div>
                  <div class="ag-name">{r['motorista']}</div>
                  <div class="ag-sub">Saída: {hora_ag} &bull; Retorno est.: {retorno} &bull; {r['num_coletas']} coletas &bull; {r['distancia_km']} km &bull; {formatar_tempo(tempo_est)}</div>
                </div>
              </div>
              <div class="timeline">
            """, unsafe_allow_html=True)

            for pi, (p, eta) in enumerate(zip(r["paradas"], etas_ag)):
                is_base = pi == 0 or pi == len(r["paradas"]) - 1
                cls = "tl-item base-item" if is_base else "tl-item"
                dist_info = f"· {r['trechos'][pi-1]['distancia_km']} km do anterior" if pi > 0 and pi - 1 < len(r["trechos"]) else ""
                icon_base = "🏠" if pi == 0 else ("🔚" if is_base else f"{pi}.")
                st.markdown(f"""
                <div class="{cls}">
                  <span class="tl-time">{eta} &nbsp; {icon_base} {p['nome'][:50]}</span>
                  {"<span class='tl-dist'>" + dist_info + "</span>" if dist_info else ""}
                </div>
                """, unsafe_allow_html=True)

            st.markdown("</div></div>", unsafe_allow_html=True)

        # Visão tabular comparativa de horários
        st.markdown('<div class="section-title">Cronograma Comparativo</div>', unsafe_allow_html=True)
        max_paradas = max(len(r["paradas"]) for r in rotas_ag)
        colunas_head = ["Parada #"] + [r["motorista"] for r in rotas_ag]
        rows_tabela = []
        for pi in range(max_paradas):
            row = {"Parada #": pi}
            for r in rotas_ag:
                etas_t = calcular_etas(r["paradas"], hora_ag, vel_ag, t_ag)
                if pi < len(r["paradas"]):
                    nome = r["paradas"][pi]["nome"][:25]
                    eta = etas_t[pi] if pi < len(etas_t) else "-"
                    row[r["motorista"]] = f"{eta} — {nome}"
                else:
                    row[r["motorista"]] = "—"
            rows_tabela.append(row)
        st.dataframe(pd.DataFrame(rows_tabela), use_container_width=True, hide_index=True)


# ══════════════════════════════════════════════════════════════════════════════
# TAB STATUS
# ══════════════════════════════════════════════════════════════════════════════
with tab_status:
    st.markdown('<div class="section-title">Status das Coletas</div>', unsafe_allow_html=True)
    st.markdown('<div class="section-subtitle">Acompanhe em tempo real o andamento de cada ponto</div>', unsafe_allow_html=True)

    if not st.session_state.pontos_coleta:
        st.info("Nenhum ponto cadastrado.")
    else:
        pontos_st = st.session_state.pontos_coleta
        status_map = st.session_state.status_pontos

        # KPIs de status
        total = len(pontos_st)
        coletados = sum(1 for i in range(total) if status_map.get(i) == "coletado")
        falhou = sum(1 for i in range(total) if status_map.get(i) == "falhou")
        reagendado = sum(1 for i in range(total) if status_map.get(i) == "reagendado")
        pendentes = total - coletados - falhou - reagendado
        pct = int(coletados / max(total, 1) * 100)

        c1, c2, c3, c4 = st.columns(4)
        with c1: st.metric("✅ Coletados", coletados)
        with c2: st.metric("⏳ Pendentes", pendentes)
        with c3: st.metric("❌ Falhou", falhou)
        with c4: st.metric("🔄 Reagendado", reagendado)

        st.markdown(f"""
        <div style="margin:8px 0 16px">
          <div style="font-size:.82rem;color:#7f8c8d;margin-bottom:4px">Progresso: <b>{pct}%</b> ({coletados}/{total})</div>
          <div class="prog-bar-outer"><div class="prog-bar-inner" style="width:{pct}%"></div></div>
        </div>
        """, unsafe_allow_html=True)

        # Filtro
        filtro_st = st.selectbox(
            "Filtrar por status",
            ["Todos", "pendente", "coletado", "falhou", "reagendado"],
            key="filtro_status",
        )

        st.divider()
        nova_status = dict(status_map)
        for p_idx, p in enumerate(pontos_st):
            st_atual = status_map.get(p_idx, "pendente")
            if filtro_st != "Todos" and st_atual != filtro_st:
                continue

            mot_idx = st.session_state.atribuicao_motorista.get(p_idx, None)
            mot_nome = ""
            mot_cor = "#95a5a6"
            if mot_idx is not None and mot_idx < len(st.session_state.motoristas):
                mot_nome = st.session_state.motoristas[mot_idx]["nome"]
                mot_cor = CORES_MOTORISTAS[mot_idx % len(CORES_MOTORISTAS)]

            sc1, sc2, sc3 = st.columns([3, 2, 2])
            with sc1:
                st_badge_cls = f"status-{st_atual}"
                icon_st = STATUS_ICONE.get(st_atual, "⏳")
                st.markdown(
                    f'<div style="display:flex;align-items:center;gap:8px;padding:6px 0">'
                    f'<span class="status-badge {st_badge_cls}">{icon_st} {st_atual}</span>'
                    f'<span style="font-weight:600;font-size:.9rem">#{p_idx+1} — {p.get("nome","")[:40]}</span>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
                if mot_nome:
                    st.markdown(
                        f'<span style="font-size:.75rem;color:{mot_cor};font-weight:600">■ {mot_nome}</span>',
                        unsafe_allow_html=True,
                    )
            with sc2:
                novo_st = st.selectbox(
                    "Status",
                    options=STATUS_PONTO,
                    index=STATUS_PONTO.index(st_atual) if st_atual in STATUS_PONTO else 0,
                    key=f"st_{p_idx}",
                    label_visibility="collapsed",
                )
                nova_status[p_idx] = novo_st
            with sc3:
                obs_st = st.text_input(
                    "Obs", value=p.get("obs", ""),
                    key=f"obs_st_{p_idx}", label_visibility="collapsed",
                    placeholder="Observação…",
                )
                if obs_st != p.get("obs", ""):
                    st.session_state.pontos_coleta[p_idx]["obs"] = obs_st

        col_sv1, col_sv2 = st.columns(2)
        with col_sv1:
            if st.button("💾 Salvar Status", type="primary", use_container_width=True):
                st.session_state.status_pontos = nova_status
                st.success("Status atualizado!")
                st.rerun()
        with col_sv2:
            if st.button("🔄 Resetar Todos para Pendente", use_container_width=True):
                st.session_state.status_pontos = {}
                st.rerun()

        # Exportar relatório de status
        if st.session_state.status_pontos:
            st.divider()
            rows_stat = []
            for p_idx, p in enumerate(pontos_st):
                mot_idx = st.session_state.atribuicao_motorista.get(p_idx, None)
                rows_stat.append({
                    "#": p_idx + 1,
                    "Local": p.get("nome", ""),
                    "Endereço": p.get("endereco", ""),
                    "Motorista": st.session_state.motoristas[mot_idx]["nome"] if mot_idx is not None and mot_idx < len(st.session_state.motoristas) else "-",
                    "Status": status_map.get(p_idx, "pendente"),
                    "Obs": p.get("obs", ""),
                })
            df_status_exp = pd.DataFrame(rows_stat)
            st.download_button(
                "📄 Exportar Status CSV",
                df_status_exp.to_csv(index=False),
                "skw_status_coletas.csv", "text/csv", use_container_width=True,
            )


# ══════════════════════════════════════════════════════════════════════════════
# TAB RELATORIO
# ══════════════════════════════════════════════════════════════════════════════
with tab_relatorio:
    st.markdown('<div class="section-title">Relatorio Gerencial SKW</div>', unsafe_allow_html=True)

    if st.session_state.pontos_coleta:
        rotas = st.session_state.rota_otimizada
        if rotas:
            dist_total = sum(r["distancia_km"] for r in rotas)
            litros_total = sum(r.get("litros_estimados", 0) for r in rotas)
            custo_total = sum(r.get("custo_combustivel", 0) for r in rotas)
            coletas_total = sum(r["num_coletas"] for r in rotas)
            tempo_total = sum(estimar_tempo_rota(
                r["distancia_km"], r["num_coletas"],
                st.session_state.velocidade_media, st.session_state.tempo_coleta_min
            ) for r in rotas)

            st.markdown(f"""
            <div class="kpi-row">
                <div class="kpi-card blue"><div class="kpi-value">{coletas_total}</div><div class="kpi-label">Coletas</div></div>
                <div class="kpi-card green"><div class="kpi-value">{len(rotas)}</div><div class="kpi-label">Motoristas</div></div>
                <div class="kpi-card orange"><div class="kpi-value">{dist_total:.1f} km</div><div class="kpi-label">Distancia</div></div>
                <div class="kpi-card red"><div class="kpi-value">R$ {custo_total:.2f}</div><div class="kpi-label">Custo</div></div>
                <div class="kpi-card purple"><div class="kpi-value">R$ {custo_total/max(coletas_total,1):.2f}</div><div class="kpi-label">Custo/Coleta</div></div>
                <div class="kpi-card teal"><div class="kpi-value">{formatar_tempo(tempo_total)}</div><div class="kpi-label">Tempo Total</div></div>
            </div>
            """, unsafe_allow_html=True)

            if len(rotas) > 1:
                st.markdown('<div class="section-title">Equilibrio</div>', unsafe_allow_html=True)
                dists = [r["distancia_km"] for r in rotas]
                eq1, eq2, eq3, eq4 = st.columns(4)
                with eq1:
                    st.metric("Desvio Padrao", f"{np.std(dists):.2f} km")
                with eq2:
                    st.metric("Razao Max/Min", f"{max(dists)/max(min(dists),0.01):.1f}x")
                with eq3:
                    st.metric("Mais Distante", f"{max(dists):.2f} km")
                with eq4:
                    st.metric("Menos Distante", f"{min(dists):.2f} km")

            df_rel = pd.DataFrame([{
                "Motorista": r["motorista"], "Coletas": r["num_coletas"],
                "Dist (km)": r["distancia_km"], "Litros": r.get("litros_estimados", 0),
                "Custo (R$)": r.get("custo_combustivel", 0),
                "Tempo": formatar_tempo(estimar_tempo_rota(
                    r["distancia_km"], r["num_coletas"],
                    st.session_state.velocidade_media, st.session_state.tempo_coleta_min)),
                "km/Coleta": round(r["distancia_km"] / max(r["num_coletas"], 1), 2),
                "Maior Trecho (km)": max((t["distancia_km"] for t in r["trechos"]), default=0),
            } for r in rotas])
            st.dataframe(df_rel, use_container_width=True, hide_index=True)

            st.markdown('<div class="section-title">Top 10 Maiores Trechos</div>', unsafe_allow_html=True)
            trechos_info = []
            for r in rotas:
                for t in r["trechos"]:
                    trechos_info.append({"Motorista": r["motorista"], "De": t["de_nome"],
                                         "Para": t["para_nome"], "Dist (km)": t["distancia_km"]})
            df_top = pd.DataFrame(trechos_info).sort_values("Dist (km)", ascending=False).head(10)
            st.dataframe(df_top, use_container_width=True, hide_index=True)
        else:
            st.info("Otimize as rotas para ver o relatorio completo.")

        if geojson_bairros:
            st.markdown('<div class="section-title">Distribuicao por Bairro</div>', unsafe_allow_html=True)
            bairros = [ponto_no_poligono(p["lat"], p["lon"], geojson_bairros)
                       for p in st.session_state.pontos_coleta]
            contagem = pd.DataFrame({"Bairro": bairros})["Bairro"].value_counts().reset_index()
            contagem.columns = ["Bairro", "Pontos"]
            st.dataframe(contagem, use_container_width=True, hide_index=True)
            st.bar_chart(contagem.set_index("Bairro"))
    else:
        st.info("Importe pontos para gerar o relatorio.")


# ══════════════════════════════════════════════════════════════════════════════
# FOOTER
# ══════════════════════════════════════════════════════════════════════════════
st.markdown(f"""
<div class="footer-pro">
    <b>SKW COLETAS v2.0</b> &bull; Sistema Profissional de Roteirização e Coletas &bull; Goiânia/GO<br>
    OR-Tools + 2-opt + Or-opt &bull; CVRP (Capacidade) &bull; ETAs por Parada &bull; Export Excel/PDF &bull; Agenda &bull; Status em Tempo Real<br>
    <span style="color:#bdc3c7">Streamlit &bull; Folium &bull; OpenPyXL &bull; GeoPy &bull; Shapely</span>
</div>
""", unsafe_allow_html=True)
